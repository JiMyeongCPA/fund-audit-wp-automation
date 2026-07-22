"""General row layout for C2_자산부채평가's detail sections.

Every section follows the same shape: a numbered caption row, a column-header
row (bottom border), then a listing area of exactly N+3 rows (N real entries
+ 3 blank buffer rows, regardless of N -- 0 real rows still shows 3 blank
rows), then a separate totals row (top+double-bottom border, yellow fill on
the total cells) summing that whole listing area.

This is a deliberate departure from copying C2 verbatim: a copied sheet's
row capacity is frozen at whatever the reference happened to need, so PBC
data that grows past it silently drops out of a SUM range (or, worse, has no
formula slot at all -- verified against the real file for 채권명세, which
has zero spare capacity). Building each block fresh from the current PBC
data's real row count removes that ceiling entirely: capacity is recomputed
every run, not carried over from history.

A block with zero real rows is hidden outright (all its rows, header through
total) -- verified against the real file: 콜론/REPO매수/매입어음/전자단기사채
are all zero this period and their rows are all row_dimensions[...].hidden,
while 채권 (non-zero) isn't.

Every row also carries, in column A, a live `=IF($G{summary_row}=0,"O","X")`
tie-out check against that category's own summary-table 소계 cell (G column),
plus a solid theme-3 fill down column B -- both copied patterns from the
real file's own detail blocks, not something we invented.
"""
from __future__ import annotations

from typing import Callable

from openpyxl.styles import Border, Color, PatternFill, Side

BUFFER_ROWS = 3
COLUMN_HEADER_BORDER = Border(bottom=Side(style="thin"))
TOTAL_ROW_BORDER = Border(top=Side(style="thin"), bottom=Side(style="double"))
COLUMN_B_FILL = PatternFill(patternType="solid", fgColor=Color(theme=3, tint=0.0))
TOTAL_CELL_FILL = PatternFill(patternType="solid", fgColor="FFFFFF00")


def write_block(
    ws,
    start_row: int,
    header_text: str,
    column_headers: list[str],
    row_writers: list[Callable],
    total_cols: list[str],
    hidden: bool = False,
    summary_row: int | None = None,
) -> dict:
    """Write one detail block starting at `start_row`.

    `row_writers[i](ws, row)` writes real data row `i`'s cells at sheet row
    `row`. `total_cols` are the column letters (e.g. ["F", "G", "H"]) that
    get a `=SUM(...)` totals formula covering the whole data+buffer range.
    Pass an empty list if the block has no totals row of its own (e.g. 채권,
    whose summary-table row does its own SUMIF directly over the block's raw
    row range instead of reading a single total cell).

    `hidden=True` marks every row in the block (header through total)
    Excel-hidden -- pass this for blocks with zero real rows.

    `summary_row`, if given, is this category's row in the summary table
    (rows 12-34); every row in the block gets a column-A tie-out check
    formula referencing that row's G (소계) cell, plus the column-B fill.

    Returns first_data_row/last_data_row (the whole data+buffer range),
    total_row (None if total_cols was empty), and next_row (where the next
    block's header goes -- one blank row after whatever this block's last
    row turned out to be).
    """
    last_col = 3 + len(column_headers) - 1

    ws.cell(start_row, 3, header_text)
    for i, header in enumerate(column_headers):
        ws.cell(start_row + 1, 3 + i, header)
    for col in range(3, last_col + 1):
        ws.cell(start_row + 1, col).border = COLUMN_HEADER_BORDER

    first_data_row = start_row + 2
    for i, writer in enumerate(row_writers):
        writer(ws, first_data_row + i)

    last_data_row = first_data_row + len(row_writers) - 1 + BUFFER_ROWS

    total_row = None
    if total_cols:
        total_row = last_data_row + 1
        for col_letter in total_cols:
            cell = ws[f"{col_letter}{total_row}"]
            cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
            cell.fill = TOTAL_CELL_FILL
        for col in range(3, last_col + 1):
            ws.cell(total_row, col).border = TOTAL_ROW_BORDER

    last_row = total_row if total_row is not None else last_data_row

    if summary_row is not None:
        for row in range(start_row, last_row + 1):
            ws.cell(row, 1, f'=IF($G{summary_row}=0,"O","X")')
            ws.cell(row, 2).fill = COLUMN_B_FILL

    if hidden:
        for row in range(start_row, last_row + 1):
            ws.row_dimensions[row].hidden = True

    next_row = last_row + 2
    return {
        "first_data_row": first_data_row,
        "last_data_row": last_data_row,
        "total_row": total_row,
        "next_row": next_row,
        "hidden": hidden,
    }

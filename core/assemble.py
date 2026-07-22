"""Assemble a workpaper from scratch: each sheet's origin is explicit --
parsed from PBC, copied from the reference workpaper, or (later) computed --
rather than loading a full template and leaving most of it untouched.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from core import (
    bosubunbae,
    c2_assemble,
    c3_income,
    c4_fee,
    c5_assemble,
    chaegwon,
    gajungpyeonggyunjwasu,
    gijunkagyeok,
    gungnaejusik,
    gungnaeyudong,
    ilbyeoljasan,
    pundbyeolmyeongse,
    pyeongga_jusik,
    pyeongga_seonmul,
    pyeongga_sooikjeunggwon,
    seonmul,
    seoljeongheji,
    settlement_sheet,
    sooikjeunggwon,
)
from core.fee_rate_source import extract_fee_rates
from core.interest_rate_source import extract_day_split, extract_interest_rates
from core.prior_year_source import (
    extract_prior_year_capital_statement,
    extract_prior_year_financials,
)
from core.sheet_copy import copy_sheet

# 주요 검토 시트의 머리말 블록(1~4행, A~L열)은 조서 본문에서 비워둔다 --
# 펀드명/작성자(Prepared by)/검토자(Reviewed by) 같은 식별성 헤더가 산출물에
# 남지 않도록. 참조 워크북에서 복사되거나(C1_정산표) 그걸 수식으로 참조하는
# (C2~C5·평가액검증) 시트 모두에 적용해야 값이 뚫고 올라오지 않는다.
_HEADER_CLEAR_SHEETS = [
    "C1_정산표",
    "C2_자산부채평가",
    "C3_수익",
    "평가액검증(주식)",
    "평가액검증(수익증권)",
    "평가액검증(선물)",
    "C4_수수료비용 등",
    "C5_F.N",
]


def _clear_header_rows(wb) -> None:
    for name in _HEADER_CLEAR_SHEETS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        for row in range(1, 5):
            for col in range(1, 13):
                ws.cell(row, col).value = None


def build_workpaper(
    gijunkagyeok_csv_path: str | Path,
    gungnaeyudong_csv_path: str | Path,
    gungnaejusik_csv_path: str | Path,
    pundbyeolmyeongse_csv_path: str | Path,
    sooikjeunggwon_csv_path: str | Path,
    seonmul_csv_path: str | Path,
    chaegwon_csv_path: str | Path,
    gajungpyeonggyunjwasu_csv_path: str | Path,
    bosubunbae_csv_path: str | Path,
    ilbyeoljasan_csv_path: str | Path,
    seoljeongheji_csv_path: str | Path,
    reference_workpaper_path: str | Path,
    output_path: str | Path,
) -> dict:
    """`reference_workpaper_path` stands in for last period's completed
    workpaper (전기 조서) -- currently the only one available is 샘플펀드's own
    당기 완료본, used as a stand-in until a real 전기 조서 exists.
    """
    df = gijunkagyeok.parse(str(gijunkagyeok_csv_path))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    gijunkagyeok.create_sheet(wb, df)

    ref_wb = openpyxl.load_workbook(str(reference_workpaper_path), data_only=False)
    missing_codes = gijunkagyeok.find_missing_account_codes(ref_wb, df)
    copy_sheet(ref_wb, gijunkagyeok.MASTER_SHEET_NAME, wb, unhide=True)

    copy_sheet(ref_wb, settlement_sheet.SHEET_NAME, wb, unhide=True)
    out_settlement_ws = wb[settlement_sheet.SHEET_NAME]

    # Parse every raw-PBC detail source now, but don't create their sheets
    # yet -- those always sit at the very end of the workbook (matching the
    # reference file's own "PBC--->>>" convention), after every sheet that
    # references them by name. Sheet name resolution doesn't care about
    # creation order, only that the sheet exists somewhere by the time Excel
    # recalculates, so building C2 first and creating these sheets last is
    # safe.
    gajungpyeonggyunjwasu_df = gajungpyeonggyunjwasu.parse(str(gajungpyeonggyunjwasu_csv_path))
    gungnaeyudong_df = gungnaeyudong.parse(str(gungnaeyudong_csv_path))
    gungnaejusik_df = gungnaejusik.parse(str(gungnaejusik_csv_path))
    pundbyeolmyeongse_df = pundbyeolmyeongse.parse(str(pundbyeolmyeongse_csv_path))
    sooikjeunggwon_df = sooikjeunggwon.parse(str(sooikjeunggwon_csv_path))
    seonmul_df = seonmul.parse(str(seonmul_csv_path))
    chaegwon_df = chaegwon.parse(str(chaegwon_csv_path))
    bosubunbae_df = bosubunbae.parse(str(bosubunbae_csv_path))
    ilbyeoljasan_df = ilbyeoljasan.parse(str(ilbyeoljasan_csv_path))
    seoljeongheji_df = seoljeongheji.parse(str(seoljeongheji_csv_path))

    # C2_자산부채평가 depends on several raw-PBC detail sheets; 외화위탁증거금
    # has no matching PBC source and is out of scope for this domestically-
    # focused fund (user's call). Assembled fresh block by block instead of
    # copied verbatim -- see core.c2_assemble.
    _, c2_block_results = c2_assemble.build_c2(
        wb, ref_wb, gungnaeyudong_df, gungnaejusik_df, sooikjeunggwon_df, seonmul_df, chaegwon_df
    )

    ref_wb_values = openpyxl.load_workbook(str(reference_workpaper_path), data_only=True)

    # C3_수익's 이자율 (D/E columns) has no PBC source -- reconstructed from
    # the reference workpaper's own values, the same "decouple acquisition
    # from use" approach as prior_year_source. C3 also needs C1_정산표 and
    # 통합기준가격대장 by name, both already copied above.
    rate_items = extract_interest_rates(ref_wb_values)
    day_split = extract_day_split(ref_wb)
    c3_income.build(wb, rate_items, day_split)

    # 평가액검증(주식/수익증권/선물): 각각 국내주식명세/수익증권명세/선물명세를
    # 기계적으로 가져와 평가금액/종가(정산가)를 검증한다. 참조하는 KRS/KRX
    # 정보시스템 시트들은 PBC 소스가 없는 순수 외부 시세 자료라 그대로 복사만
    # 하는데, 시트 배치는 원본 조서 순서(평가액검증 시트들이 먼저)를 따르고
    # KRS/KRX 자체는 맨 뒤로 보낸다 -- 아래쪽에서 한 번에 생성.
    pyeongga_jusik.build(wb, len(gungnaejusik_df))
    pyeongga_sooikjeunggwon.build(wb, len(sooikjeunggwon_df))
    pyeongga_seonmul.build(wb, len(seonmul_df))

    # C4_수수료비용 등: PL(실제 장부)보수금액 대 재계산(연평균 순자산x보수율)
    # 금액을 비교하는 분석적 검토. 보수율 자체는 PBC 소스가 없는 고정 계약
    # 조건이라 fee_rate_source에서 별도 기초자료로 재구성해 가져온다 (이자율과
    # 같은 논리). 일별 순자산 시리즈는 이미 만든 '가중평균좌수'를 그대로 쓴다.
    fee_rate_items = extract_fee_rates(ref_wb_values)
    c4_fee.build(wb, fee_rate_items, len(gajungpyeonggyunjwasu_df))

    # C5_F.N: 주석사항 검토. c5_assemble.py의 도크스트링 참고.
    _, c5_needs_review = c5_assemble.build_c5(
        wb,
        fee_rate_items,
        c2_block_results["거래소주식"]["first_data_row"],
        len(gungnaejusik_df),
        c2_block_results["채권"]["first_data_row"],
        len(chaegwon_df),
        c2_block_results["수익증권"]["first_data_row"],
        len(sooikjeunggwon_df),
        len(seonmul_df),
    )

    # 재무제표검증: C1_정산표 항목별로 재무상태표 형식에 맞춰 재배열하고 원본
    # 재무제표(전기 비교표시분 포함)의 리터럴 수치와 대사하는 시트. 수식은
    # 전부 'C1_정산표' 하나만 참조하고(이미 위에서 copy_sheet됨), 나머지 값은
    # PBC로 재구성할 수 없는 외부 재무제표 원본 수치라 KRS/KRX와 같은 성격 --
    # 그대로 복사만 한다.
    FINANCIAL_STATEMENT_CHECK_SHEET_NAME = "재무제표검증"
    copy_sheet(ref_wb, FINANCIAL_STATEMENT_CHECK_SHEET_NAME, wb, unhide=True)

    # 전기 데이터는 조서 자체가 아니라, 그 조서의 C1_정산표 전기 컬럼에서 재구성한
    # 별도 소스(prior_year_source)에서 가져온다 -- 진짜 전기 감사보고서/DSD를
    # 읽는 문제는 아직 미해결이라, 지금은 같은 파일 안에서 왕복하는 형태.
    prior_year_items = extract_prior_year_financials(ref_wb_values)
    prior_year_capital_rows = extract_prior_year_capital_statement(ref_wb_values)

    drifted_prior_year_rows = settlement_sheet.apply_prior_year_financials(
        out_settlement_ws, ref_wb_values[settlement_sheet.SHEET_NAME], prior_year_items
    )
    settlement_sheet.apply_prior_year_capital_statement(
        out_settlement_ws, prior_year_capital_rows
    )

    # 당기(G/H) 쪽은 이번 실행에서 바꾸지 않았으니, 기존 조서의 캐시된 당기 값을
    # 그대로 공시여부 판단에 쓴다. 전기(I/J) 쪽은 방금 새로 쓴 값을 쓴다.
    settlement_sheet.recompute_disclosure_hiding(
        out_settlement_ws, ref_wb_values[settlement_sheet.SHEET_NAME]
    )

    # raw-PBC detail sheets go last, matching the reference file's own
    # "PBC--->>>" convention (everything above referenced them by name already).
    gajungpyeonggyunjwasu.create_sheet(wb, gajungpyeonggyunjwasu_df)
    gungnaeyudong.create_sheet(wb, gungnaeyudong_df)
    gungnaejusik.create_sheet(wb, gungnaejusik_df)
    pundbyeolmyeongse.create_sheet(wb, pundbyeolmyeongse_df)
    sooikjeunggwon.create_sheet(wb, sooikjeunggwon_df)
    seonmul.create_sheet(wb, seonmul_df)
    chaegwon.create_sheet(wb, chaegwon_df)
    bosubunbae.create_sheet(wb, bosubunbae_df)
    ilbyeoljasan.create_sheet(wb, ilbyeoljasan_df)
    seoljeongheji.create_sheet(wb, seoljeongheji_df)

    # KRS/KRX 정보시스템 시트들(평가액검증 시트들이 참조하는 외부 시세 자료)은
    # 맨 뒤로 보낸다 -- 원본 조서에서는 평가액검증 시트들 사이사이에 숨겨진 채
    # 끼어 있었지만(KRX정보시스템_ETF/코스피200은 실제로 hidden), 여기서는 모두
    # 보이게 두고 워크북 가장 마지막에 모아둔다.
    copy_sheet(ref_wb, pyeongga_jusik.KRS_SHEET_NAME, wb, unhide=True)
    copy_sheet(ref_wb, pyeongga_sooikjeunggwon.KRX_SHEET_NAME, wb, unhide=True)
    copy_sheet(ref_wb, pyeongga_seonmul.KRX_SHEET_NAME, wb, unhide=True)

    # Captured for the review UI's "click block -> see original PBC file/row"
    # feature (core.origin_map) -- build_workpaper already has every raw path
    # and parsed df in scope, so this is just retaining what would otherwise
    # be discarded, not new parsing work.
    raw_csv_paths = {
        gijunkagyeok.SHEET_NAME: str(gijunkagyeok_csv_path),
        gungnaeyudong.SHEET_NAME: str(gungnaeyudong_csv_path),
        gungnaejusik.SHEET_NAME: str(gungnaejusik_csv_path),
        pundbyeolmyeongse.SHEET_NAME: str(pundbyeolmyeongse_csv_path),
        sooikjeunggwon.SHEET_NAME: str(sooikjeunggwon_csv_path),
        seonmul.SHEET_NAME: str(seonmul_csv_path),
        chaegwon.SHEET_NAME: str(chaegwon_csv_path),
        gajungpyeonggyunjwasu.SHEET_NAME: str(gajungpyeonggyunjwasu_csv_path),
        bosubunbae.SHEET_NAME: str(bosubunbae_csv_path),
        ilbyeoljasan.SHEET_NAME: str(ilbyeoljasan_csv_path),
        seoljeongheji.SHEET_NAME: str(seoljeongheji_csv_path),
    }
    raw_row_counts = {
        gungnaejusik.SHEET_NAME: len(gungnaejusik_df),
        chaegwon.SHEET_NAME: len(chaegwon_df),
        sooikjeunggwon.SHEET_NAME: len(sooikjeunggwon_df),
    }

    _clear_header_rows(wb)

    # openpyxl writes formulas without cached results, so every formula cell
    # (and any =A=B reconciliation check built on one) would show a stale/
    # blank value until the user manually forces a recalc -- this flag tells
    # Excel to fully recalculate everything the moment the file is opened.
    wb.calculation.fullCalcOnLoad = True

    wb.save(str(output_path))
    return {
        "missing_account_codes": missing_codes,
        "drifted_prior_year_rows": drifted_prior_year_rows,
        "c2_block_results": c2_block_results,
        "raw_csv_paths": raw_csv_paths,
        "raw_row_counts": raw_row_counts,
        "needs_review": c5_needs_review,
    }

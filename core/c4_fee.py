"""C4_수수료비용 등: "수수료비용 검토" -- an analytical review comparing each
fee category's PL(실제 장부)보수금액 against a recalculated estimate
(연평균 순자산총액 x 보수율), concluding "적정" if the total difference is
under 1% (note: a tighter threshold than C3_수익's 5%).

Two sections:

1. 수수료 비용 (rows 6-13): 4 categories (자산운용회사/판매회사/수탁회사/
   사무관리회사), each comparing D (PL보수금액, from 'C1_정산표'!G110:G113,
   already-copied) against E (재계산금액, pulled from section 2's row-387
   averages).
2. 수수료비용 재계산 (rows 19-387): a 365-day series -- 기준일/순자산총액
   pulled mechanically from '가중평균좌수' (already built from PBC data),
   then 4 columns (C-F, one per fee category) computing 순자산총액 x 그
   카테고리의 보수율 for that day. Row 387 averages each of those 365 daily
   figures -- since 보수율 is an *annual* rate and 순자산총액 varies day to
   day, AVERAGE(daily NAV x annual rate) over the year is mathematically the
   same as (연평균 순자산총액) x (연간 보수율), which is the real economic
   recalculation of an annual fee that's actually charged against a moving
   balance. This is why row 387 averages rather than sums.

Row 388 is a separate sanity check (B388 = B386 (가중평균좌수's last day's
순자산) equals 'C1_정산표'!H65 (자본 총계)) -- confirms 가중평균좌수 and
C1_정산표 agree on the year-end net asset total; not part of the fee
comparison itself, copied because it's a real check present in the file.

보수율 (C column, rows 9-12) has no PBC source -- see fee_rate_source.py for
why and how it's reconstructed as standalone base data instead of hardcoded
here directly.
"""
from __future__ import annotations

from openpyxl.styles import Border, Color, PatternFill, Side

from core.gajungpyeonggyunjwasu import SHEET_NAME as GAJUNGPYEONGGYUNJWASU_SHEET_NAME
from core.settlement_sheet import SHEET_NAME as SETTLEMENT_SHEET_NAME

SHEET_NAME = "C4_수수료비용 등"

FIRST_DAILY_ROW = 22
GAJUNGPYEONGGYUNJWASU_FIRST_DATA_ROW = 3

# (row, C1_정산표 row this category's PL보수금액 comes from, recalc column in section 2)
_CATEGORY_CONFIG = [
    (9, 110, "C"),  # 자산운용회사 / 집합투자보수
    (10, 111, "D"),  # 판매회사 / 판매보수
    (11, 112, "E"),  # 수탁회사 / 신탁보수
    (12, 113, "F"),  # 사무관리회사 / 사무관리보수
]

AMOUNT_FMT = "#,##0"
RATE_FMT = "0.000%"
RATIO_FMT = "0.0%"

_ACTIVE_FILL = PatternFill("solid", fgColor=Color(theme=5, tint=0.7999816888943144))
_BOX_BORDER = Border(top=Side(style="thin"), bottom=Side(style="thin"))
_BOTTOM_BORDER = Border(bottom=Side(style="thin"))


def build(wb, fee_rate_items, num_days: int):
    ws = wb.create_sheet(SHEET_NAME)
    g = GAJUNGPYEONGGYUNJWASU_SHEET_NAME
    s = SETTLEMENT_SHEET_NAME
    summary_row = FIRST_DAILY_ROW + num_days

    ws["A1"] = f"='{s}'!B1"
    ws["A2"] = f"='{s}'!B2"
    ws["G2"] = "Prepared by : "
    ws["H2"] = f"='{s}'!H2"
    ws["A3"] = "수수료비용 검토"
    ws["G3"] = "Reviewed by : "

    ws["A6"] = "1. 수수료 비용"
    ws["C6"] = "결론"
    ws["D6"] = '=IF(G13<0.01,"적정","부적정")'
    ws["C6"].fill = _ACTIVE_FILL
    ws["D6"].fill = _ACTIVE_FILL

    headers = ["구 분", "지급처", "보수율", "PL보수금액", "재계산금액", "차이", "차이율", "비고"]
    for col, header in zip("ABCDEFGH", headers):
        ws[f"{col}8"] = header
        ws[f"{col}8"].border = _BOX_BORDER

    for item, (row, settlement_row, recalc_col) in zip(fee_rate_items, _CATEGORY_CONFIG):
        ws[f"A{row}"] = item.label
        ws[f"B{row}"] = item.counterparty
        ws[f"C{row}"] = item.rate
        ws[f"D{row}"] = f"='{s}'!G{settlement_row}"
        ws[f"E{row}"] = f"={recalc_col}{summary_row}"
        ws[f"F{row}"] = f"=D{row}-E{row}"
        ws[f"G{row}"] = f"=F{row}/E{row}"
        ws[f"H{row}"] = "Minor pass"

        ws[f"C{row}"].number_format = RATE_FMT
        for col in "DEF":
            ws[f"{col}{row}"].number_format = AMOUNT_FMT
        ws[f"G{row}"].number_format = RATIO_FMT
        ws[f"G{row}"].fill = _ACTIVE_FILL
        for col in "ABCDEFGH":
            ws[f"{col}{row}"].border = _BOX_BORDER

    ws.merge_cells("A13:B13")
    ws["A13"] = "합 계"
    ws["D13"] = "=SUM(D9:D12)"
    ws["E13"] = "=SUM(E9:E12)"
    ws["F13"] = "=D13-E13"
    ws["G13"] = "=F13/E13"
    for col in "DEF":
        ws[f"{col}13"].number_format = AMOUNT_FMT
    ws["G13"].number_format = RATIO_FMT
    ws["G13"].fill = _ACTIVE_FILL
    for col in "ABCDEFGH":
        ws[f"{col}13"].border = _BOTTOM_BORDER

    ws["A19"] = "2. 수수료비용 재계산"

    ws.merge_cells("A20:A21")
    ws.merge_cells("B20:B21")
    ws["A20"] = "기준일 "
    ws["B20"] = "순자산총액"
    ws["C20"] = "운용보수"
    ws["D20"] = "판매보수"
    ws["E20"] = "수탁보수"
    ws["F20"] = "사무관리보수"
    for col in "ABCDEF":
        ws[f"{col}20"].border = Border(top=Side(style="thin"))
    ws["C21"] = "=C9"
    ws["D21"] = "=C10"
    ws["E21"] = "=C11"
    ws["F21"] = "=C12"
    for col in "ABCDEF":
        ws[f"{col}21"].border = Border(bottom=Side(style="thin"))
    for col in "CDEF":
        ws[f"{col}21"].number_format = RATE_FMT

    for i in range(num_days):
        row = FIRST_DAILY_ROW + i
        src_row = GAJUNGPYEONGGYUNJWASU_FIRST_DATA_ROW + i
        ws[f"A{row}"] = f"='{g}'!B{src_row}"
        ws[f"B{row}"] = f"='{g}'!D{src_row}"
        for col in "CDEF":
            ws[f"{col}{row}"] = f"=B{row}*{col}$21"

    last_daily_row = summary_row - 1
    ws[f"B{summary_row}"] = f"=SUM(AVERAGE(B{FIRST_DAILY_ROW}:B{last_daily_row}))"
    for col in "CDEF":
        ws[f"{col}{summary_row}"] = f"=AVERAGE({col}{FIRST_DAILY_ROW}:{col}{last_daily_row})"

    ws[f"B{summary_row + 1}"] = f"=B{last_daily_row}='{s}'!H65"

    return ws

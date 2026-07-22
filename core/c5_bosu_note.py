"""C5_F.N > "1. 보수": 주석 disclosure of 투자신탁보수 -- 당기/전기 보수 and
미지급보수 for each of the 4 fee counterparties, plus a 보수율 sub-table.

Nearly the entire section is a mechanical restatement of data that's already
fully wired elsewhere in this workbook:

- 당기 보수 (rows 110-113 of 'C1_정산표'!G) -- the same figures C4_수수료비용
  등 compares against its own recalculation.
- 전기 보수 (rows 110-113 of 'C1_정산표'!I) -- already reconstructed by
  prior_year_source and written into C1_정산표.
- 당기 미지급보수 (rows 56-59 of 'C1_정산표'!N) -- live formulas into
  '통합기준가격대장(결산후)' (already copied verbatim from the reference).
- 전기 미지급보수 (rows 56-59 of 'C1_정산표'!O) -- literal prior-year figures
  that our prior-year write-back never touches (that logic only overwrites
  C1_정산표's I/J columns, not M/N/O), so they're still exactly what the
  reference had.
- 지급처/보수율 -- fee_rate_source.py's FeeRateItem list, the same 4 items
  already used to build C4_수수료비용 등.

One real quirk fixed here: the real file hardcodes row 1006 (당기 보수,
수탁회사, in thousands won) as `=ROUND(N1006/1000,0)-1` -- a manual "-1"
patch so SUM(E1004:E1007) foots to ROUND(총계/1000,0) despite each line
being rounded independently (4 lines each individually rounded don't always
sum to the same result as rounding their total). Rather than guessing which
specific line the original preparer chose to absorb that difference, the
last line (사무관리회사) is written as a plug -- 반올림된 합계 빼기 앞 3개
반올림값의 합 -- so it reconciles automatically for any input instead of
needing another hardcoded adjustment if a future period rounds differently.
Verified this reproduces the *same* totals (E1008/F1008/G1008/H1008) as the
original for this fund's real data; it just assigns the ₩1천원 rounding
residual to a different (still immaterial) line.

Borders/fills below are copied from the real file's own styling (theme
colors read directly off its cells): header rows use theme 8, the 합계 row
uses theme 4 (same "computed total" highlight used elsewhere in this
project), and the tie-out check row uses theme 0. A thin grid box outlines
every row of both tables, with a medium rule down column C running the
full height of the section (title through the last row) -- also copied
directly from the real file.

One deliberate departure from the real file: the two narrative paragraphs
((2) 보수산출기간 and (3) 보수율 설명) are plain unwrapped single cells there,
spilling visually across empty neighboring cells -- fine on screen, but ugly
and hard to read once neighboring cells actually have content (as ours will,
once earlier notes are built). Here they're merged across C:H with
wrap_text on and an explicit row height, so the paragraph actually wraps
inside its own box instead of overflowing.

The real file has a "약관 확인" note next to the rate sub-table (column O,
next to its hidden L:Q master-data columns) -- dropped here rather than
carried over. It's not a formula or a data reference, just a preparer's own
review tickmark ("verified against the actual trust deed"); it doesn't
belong in the visible disclosure layer this module builds.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side

from core.c5_styles import run_left_rule

SETTLEMENT_SHEET_NAME = "C1_정산표"

# (당기/전기 보수 row in C1_정산표, 당기/전기 미지급보수 row in C1_정산표)
_CATEGORY_SETTLEMENT_ROWS = [110, 111, 112, 113]
_UNPAID_SETTLEMENT_ROWS = [56, 57, 58, 59]

AMOUNT_FMT = r"#,##0\ ;\(#,##0\);\-\ ;@"
RATE_FMT = "0.000%"  # 2 decimals rounds 1e-05 (0.001%) to 0.00%, hiding it entirely

_BOLD = Font(bold=True)
_WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")

_HEADER_FILL = PatternFill("solid", fgColor=Color(theme=8, tint=-0.249977111117893))
_TOTAL_FILL = PatternFill("solid", fgColor=Color(theme=4, tint=0.7999816888943144))
_CHECK_FILL = PatternFill("solid", fgColor=Color(theme=0, tint=-0.0499893185216834))

_THIN = Side(style="thin")
_MEDIUM = Side(style="medium")


def _box_border(is_left_edge: bool) -> Border:
    return Border(
        top=_THIN, bottom=_THIN, right=_THIN, left=_MEDIUM if is_left_edge else _THIN
    )


def _left_rule(ws, row: int) -> None:
    """The real file runs a medium vertical rule down column C for the
    entire section, title row through the last row -- not just the table
    rows."""
    cell = ws.cell(row, 3)
    cell.border = Border(left=_MEDIUM, top=cell.border.top, bottom=cell.border.bottom, right=cell.border.right)


def build(ws, fee_rate_items, start_row: int = 1000) -> int:
    """Writes the "1. 보수" note starting at `start_row` (1000 in the real
    file). Returns the next free row after this section."""
    s = SETTLEMENT_SHEET_NAME
    r = start_row

    ws.cell(r, 3, "1. 보수").font = _BOLD
    ws.cell(r + 1, 3, "(1) 당기와 전기 중 투자신탁보수의 내용은 다음과 같습니다.")
    _left_rule(ws, r)
    _left_rule(ws, r + 1)

    header_row = r + 2
    ws.cell(header_row, 3, "구 분").font = _BOLD
    ws.cell(header_row, 4, "지급처").font = _BOLD
    ws.cell(header_row, 5, "당기").font = _BOLD
    ws.cell(header_row, 7, "전기").font = _BOLD
    ws.merge_cells(start_row=header_row, start_column=3, end_row=header_row + 1, end_column=3)
    ws.merge_cells(start_row=header_row, start_column=4, end_row=header_row + 1, end_column=4)
    ws.merge_cells(start_row=header_row, start_column=5, end_row=header_row, end_column=6)
    ws.merge_cells(start_row=header_row, start_column=7, end_row=header_row, end_column=8)

    subheader_row = header_row + 1
    ws.cell(subheader_row, 5, "보수")
    ws.cell(subheader_row, 6, "보수미지급금")
    ws.cell(subheader_row, 7, "보수")
    ws.cell(subheader_row, 8, "보수미지급금")

    for row in (header_row, subheader_row):
        for col in range(3, 9):
            cell = ws.cell(row, col)
            cell.border = _box_border(col == 3)
            cell.fill = _HEADER_FILL

    first_data_row = subheader_row + 1
    for i, item in enumerate(fee_rate_items):
        row = first_data_row + i
        settlement_row = _CATEGORY_SETTLEMENT_ROWS[i]
        unpaid_row = _UNPAID_SETTLEMENT_ROWS[i]

        ws.cell(row, 3, item.label)
        ws.cell(row, 4, item.counterparty)
        if i < len(fee_rate_items) - 1:
            ws.cell(row, 5, f"=ROUND('{s}'!G{settlement_row}/1000,0)")
            ws.cell(row, 6, f"=ROUND('{s}'!N{unpaid_row}/1000,0)")
            ws.cell(row, 7, f"=ROUND('{s}'!I{settlement_row}/1000,0)")
            ws.cell(row, 8, f"=ROUND('{s}'!O{unpaid_row}/1000,0)")
        else:
            # plug row: absorbs whatever residual independent rounding
            # leaves behind, so the displayed lines always foot exactly.
            prev_rows = range(first_data_row, row)
            ws.cell(row, 5, (
                f"=ROUND(SUM('{s}'!G{_CATEGORY_SETTLEMENT_ROWS[0]}:G{_CATEGORY_SETTLEMENT_ROWS[-1]})/1000,0)"
                f"-SUM(E{prev_rows.start}:E{prev_rows.stop - 1})"
            ))
            ws.cell(row, 6, (
                f"=ROUND(SUM('{s}'!N{_UNPAID_SETTLEMENT_ROWS[0]}:N{_UNPAID_SETTLEMENT_ROWS[-1]})/1000,0)"
                f"-SUM(F{prev_rows.start}:F{prev_rows.stop - 1})"
            ))
            ws.cell(row, 7, (
                f"=ROUND(SUM('{s}'!I{_CATEGORY_SETTLEMENT_ROWS[0]}:I{_CATEGORY_SETTLEMENT_ROWS[-1]})/1000,0)"
                f"-SUM(G{prev_rows.start}:G{prev_rows.stop - 1})"
            ))
            ws.cell(row, 8, (
                f"=ROUND(SUM('{s}'!O{_UNPAID_SETTLEMENT_ROWS[0]}:O{_UNPAID_SETTLEMENT_ROWS[-1]})/1000,0)"
                f"-SUM(H{prev_rows.start}:H{prev_rows.stop - 1})"
            ))

        for col in range(3, 9):
            ws.cell(row, col).border = _box_border(col == 3)
        for col in (5, 6, 7, 8):
            ws.cell(row, col).number_format = AMOUNT_FMT

    total_row = first_data_row + len(fee_rate_items)
    ws.cell(total_row, 3, "합 계")
    for col in (5, 6, 7, 8):
        col_letter = "EFGH"[col - 5]
        ws.cell(total_row, col, f"=SUM({col_letter}{first_data_row}:{col_letter}{total_row - 1})")
        ws.cell(total_row, col).number_format = AMOUNT_FMT
    for col in range(3, 9):
        ws.cell(total_row, col).border = _box_border(col == 3)
        ws.cell(total_row, col).fill = _TOTAL_FILL

    check_row = total_row + 1
    settlement_first, settlement_last = _CATEGORY_SETTLEMENT_ROWS[0], _CATEGORY_SETTLEMENT_ROWS[-1]
    unpaid_first, unpaid_last = _UNPAID_SETTLEMENT_ROWS[0], _UNPAID_SETTLEMENT_ROWS[-1]
    ws.cell(check_row, 5, (
        f"=E{total_row}=ROUND(SUM('{s}'!G{settlement_first}:G{settlement_last})/1000,0)"
    ))
    ws.cell(check_row, 6, (
        f"=F{total_row}=ROUND(SUM('{s}'!N{unpaid_first}:N{unpaid_last})/1000,0)"
    ))
    ws.cell(check_row, 5).fill = _CHECK_FILL
    ws.cell(check_row, 6).fill = _CHECK_FILL
    _left_rule(ws, check_row)

    narrative_row_1 = check_row + 1
    ws.merge_cells(start_row=narrative_row_1, start_column=3, end_row=narrative_row_1, end_column=8)
    narrative_cell_1 = ws.cell(narrative_row_1, 3, (
        "(2)  보수산출기간 및 인출일\n"
        "투자신탁보수의 계산기간은 투자신탁의 최초설정일로부터 매 3개월간으로 하며 보수계산기간 중 투자신탁보수를 "
        "매일 재무상태표상에 계상하고 보수계산기간의 종료시, 투자신탁의 일부해지 또는 전부해지시에 자산운용회사의 "
        "지시에 따라 수탁회사가 투자신탁 재산에서 인출하고 있습니다."
    ))
    narrative_cell_1.alignment = _WRAP_TOP
    ws.row_dimensions[narrative_row_1].height = 75
    _left_rule(ws, narrative_row_1)

    narrative_row_2 = narrative_row_1 + 2
    ws.merge_cells(start_row=narrative_row_2, start_column=3, end_row=narrative_row_2, end_column=8)
    narrative_cell_2 = ws.cell(narrative_row_2, 3, (
        "(3)  투자신탁은 신탁약관 제39조에 의거 다음의 보수율에 투자신탁재산의 순자산총액의 평균잔액"
        "(매일의 투자신탁 순자산총액을 보수계산기간의 초일부터 보수계산당일까지 누적하여 합한 금액을 "
        "보수계산기간의 일수로 나눈 금액)을 곱한 금액을 매 3개월마다 투자신탁보수로 지급하고 있습니다."
    ))
    narrative_cell_2.alignment = _WRAP_TOP
    ws.row_dimensions[narrative_row_2].height = 60
    _left_rule(ws, narrative_row_2)
    for row in range(narrative_row_1 + 1, narrative_row_2):
        _left_rule(ws, row)

    rate_header_row = narrative_row_2 + 2
    ws.cell(rate_header_row, 3, "구 분").font = _BOLD
    ws.cell(rate_header_row, 4, "지급처")
    ws.cell(rate_header_row, 5, "보수율")
    for col in (3, 4, 5):
        cell = ws.cell(rate_header_row, col)
        cell.border = _box_border(col == 3)
        cell.fill = _HEADER_FILL
    for row in range(narrative_row_2 + 1, rate_header_row):
        _left_rule(ws, row)

    rate_first_row = rate_header_row + 1
    for i, item in enumerate(fee_rate_items):
        row = rate_first_row + i
        ws.cell(row, 3, item.label)
        ws.cell(row, 4, item.counterparty)
        ws.cell(row, 5, item.rate)
        ws.cell(row, 5).number_format = RATE_FMT
        for col in (3, 4, 5):
            ws.cell(row, col).border = _box_border(col == 3)

    rate_last_row = rate_first_row + len(fee_rate_items) - 1

    run_left_rule(ws, start_row, rate_last_row)

    return rate_last_row + 2

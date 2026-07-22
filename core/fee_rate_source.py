"""Build a standalone "수수료율 기초자료" workbook out of C4_수수료비용 등's
own 구분/지급처/보수율 (A/B/C columns, rows 9-12) -- a stand-in for "look up
the real fee schedule in the fund's prospectus/집합투자규약" until that's
actually automatable, the same "decouple acquisition from use" approach as
prior_year_source.py and interest_rate_source.py.

These 4 rates (운용보수/판매보수/신탁보수/사무관리보수) are fixed contractual
terms, not something any PBC transaction file would contain -- there's no
daily activity that would let you derive "0.001% per annum" from first
principles, unlike an average balance or a recalculated income figure.
"""
from __future__ import annotations

from dataclasses import dataclass

import openpyxl

CATEGORY_ROWS = [9, 10, 11, 12]  # 자산운용회사/판매회사/수탁회사/사무관리회사
LABEL_COL = 1  # A열 구분
COUNTERPARTY_COL = 2  # B열 지급처
RATE_COL = 3  # C열 보수율


@dataclass
class FeeRateItem:
    label: str
    counterparty: str
    rate: float


def extract_fee_rates(reference_wb_values) -> list[FeeRateItem]:
    """`reference_wb_values` must be data_only=True."""
    ws = reference_wb_values["C4_수수료비용 등"]
    items = []
    for row in CATEGORY_ROWS:
        items.append(
            FeeRateItem(
                label=ws.cell(row, LABEL_COL).value,
                counterparty=ws.cell(row, COUNTERPARTY_COL).value,
                rate=ws.cell(row, RATE_COL).value,
            )
        )
    return items


def write_fee_rate_workbook(items: list[FeeRateItem], output_path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "수수료율"

    ws.append(["구분", "지급처", "보수율"])
    for item in items:
        ws.append([item.label, item.counterparty, item.rate])
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12

    wb.save(str(output_path))

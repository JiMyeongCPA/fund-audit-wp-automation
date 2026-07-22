"""평가액검증(주식): row-by-row valuation check for every domestic stock
position -- columns 1-7 (종목/종목명/상품구분/수량/취득가액/종가/평가금액) are
pulled mechanically from '국내주식명세' (already built by gungnaejusik.py),
then three checks are added:

- 평가금액정확성 (H): 수량 x 종가 == 평가금액 (a boolean formula).
- 종가검증 A (I): looks up the same stock's 종가 in 'KRS정보시스템_주식' (a
  plain external market-data dump, copied verbatim from the reference
  workpaper -- no PBC source exists for it) via INDEX/MATCH on 종목명.
  Wrapped in IFERROR, falling back to F (그 종목의 자기 종가) when the
  lookup fails -- verified against the real file: 2 of 204 stocks
  ('한온시스템[유상]', '영풍[배당]', temporary tickers for shares from a
  rights offering / stock dividend that trade under a bracket-suffixed name
  for a short window before merging into the regular listing) have no
  matching line in KRS정보시스템_주식 at all, by name or by code. The real
  file handles this by manually overwriting just those two cells with a
  hardcoded literal equal to F -- i.e. "checked by hand, matches the
  recorded price". The IFERROR fallback reproduces the same result as a
  live formula instead, so it self-heals for any future occurrence of this
  naming pattern rather than needing another manual literal.
- 종가검증 B (J): TRUE/FALSE text comparing F (국내주식명세's 종가) against I
  (KRS정보시스템_주식's 종가) -- catches a stale/incorrect 종가 in the PBC
  data itself.

Row 9's FALSE-count summary is dynamic, sized to however many stock rows
exist (`num_stock_rows`), unlike the reference workpaper's own version:
that one hardcodes 'H12:H82' / 'I12:I82' -- a range that only covers the
first 71 of what are, for 샘플펀드, actually 204 stock rows (12-215), so 133
rows are silently excluded from the FALSE count. Its second formula is also
a copy-paste bug: it counts zeros in I12:I82 (numeric 종가 A values, which
are never really 0) instead of counting "FALSE" text results in J (종가검증
B, the column the label and adjacent count cell are actually meant to
summarize). Both are fixed here: the range always spans the real row count,
and J9 counts "FALSE" text in the J column directly.
"""
from __future__ import annotations

from openpyxl.styles import Border, PatternFill, Side

from core.gungnaejusik import SHEET_NAME as GUNGNAEJUSIK_SHEET_NAME

SHEET_NAME = "평가액검증(주식)"
KRS_SHEET_NAME = "KRS정보시스템_주식"

FIRST_DATA_ROW = 12
SOURCE_FIRST_DATA_ROW = 3  # 국내주식명세's own first data row

ACCOUNTING_FMT = r'_-* #,##0_-;\-* #,##0_-;_-* "-"_-;_-@_-'

_COUNT_FILL = PatternFill("solid", fgColor="FFFFCCFF")
_TEST_LABEL_FILL = PatternFill("solid", fgColor="FFFFFF00")
_TEST_GROUP_BORDER = Border(top=Side(style="medium"), bottom=Side(style="thin"))
_TEST_LABEL_BORDER = Border(top=Side(style="thin"), bottom=Side(style="medium"))

_COLUMN_LABELS = [
    "종목            ",
    "종목명          ",
    "상품구분        ",
    "수량            ",
    "취득가액        ",
    "종가            ",
    "평가금액        ",
]


def build(wb, num_stock_rows: int):
    ws = wb.create_sheet(SHEET_NAME)
    g = GUNGNAEJUSIK_SHEET_NAME

    ws["B1"] = "='C1_정산표'!B1"
    ws["B2"] = "='C1_정산표'!B2"
    ws["G2"] = "Prepared by : "
    ws["H2"] = "='C1_정산표'!H2"
    ws["B3"] = "평가액검증"
    ws["G3"] = "Reviewed by : "
    ws["H3"] = " "

    last_row = FIRST_DATA_ROW + num_stock_rows - 1

    ws["G9"] = "FALSE의 개수=>"
    ws["H9"] = f"=COUNTIF(H{FIRST_DATA_ROW}:H{last_row},0)"
    ws["J9"] = f'=COUNTIF(J{FIRST_DATA_ROW}:J{last_row},"FALSE")'
    ws["H9"].fill = _COUNT_FILL
    ws["J9"].fill = _COUNT_FILL

    for col, label in zip("ABCDEFG", [f"Column{i}" for i in range(1, 8)]):
        ws[f"{col}10"] = label
    ws["H10"] = "Test 1"
    ws["I10"] = "Test 2"
    ws.merge_cells("I10:J10")
    for col in "HIJ":
        ws[f"{col}10"].border = _TEST_GROUP_BORDER

    for col, label in zip("ABCDEFG", _COLUMN_LABELS):
        ws[f"{col}11"] = label
    ws["H11"] = "평가금액정확성"
    ws["I11"] = "종가검증 A"
    ws["J11"] = "종가검증 B"
    for col in "HIJ":
        ws[f"{col}11"].fill = _TEST_LABEL_FILL
        ws[f"{col}11"].border = _TEST_LABEL_BORDER

    for i in range(num_stock_rows):
        row = FIRST_DATA_ROW + i
        src_row = SOURCE_FIRST_DATA_ROW + i
        for col, src_col in zip("ABCDEFG", "ABCDEFG"):
            ws[f"{col}{row}"] = f"='{g}'!{src_col}{src_row}"
        ws[f"H{row}"] = f"=G{row}=D{row}*F{row}"
        ws[f"I{row}"] = (
            f"=IFERROR(INDEX({KRS_SHEET_NAME}!$A$1:$N$3000,"
            f"MATCH('{SHEET_NAME}'!B{row},{KRS_SHEET_NAME}!$B:$B,0),"
            f"MATCH({KRS_SHEET_NAME}!$E$1,{KRS_SHEET_NAME}!$1:$1,0)),F{row})"
        )
        ws[f"J{row}"] = f'=IF(D{row}=0,"TRUE",IF(F{row}=I{row},"TRUE","FALSE"))'
        for col in "GHIJ":
            ws[f"{col}{row}"].number_format = ACCOUNTING_FMT

    return ws

"""평가액검증(선물): row-by-row valuation check for every futures position
held by each counterparty, pulling columns A-M mechanically from '선물명세'
(already built by seonmul.py) except C (종목코드), which is derived from D
(표준코드) via `MID(D,4,8)` in the real file -- not a straight column pull.

- 평가금액정확성 (N): 계약수 x 정산가 x 250,000 == 평가금액. The 250,000
  multiplier is the KOSPI200 futures contract's won-per-index-point value,
  not an arbitrary constant.
- 정산가검증 A (O): looks up this instrument's 종가 in
  'KRX정보시스템_코스피200' via INDEX/MATCH on 종목코드 (again a
  self-referencing $C$1 trick: KRX정보시스템_코스피200!C1 holds the literal
  '종가', matching its own C-column header, so MATCH resolves back to
  column C even though the row-level column being checked is 정산가/H).
- 정산가검증 B (P): TRUE/FALSE comparing H (this fund's own 정산가) against
  O, skipping straight to "TRUE" when H is 0 -- same guard shape as
  pyeongga_jusik's D=0 guard: a position with no live settlement price has
  nothing meaningful to compare against a market price, so it shouldn't be
  flagged FALSE just because it's degenerate (e.g. a contract already
  closed out).

Two things fixed relative to the real file, both verified directly against
it (not assumed):

1. A (매매처 코드) is a hardcoded literal ('\\t000012', etc.) in the real
   file even though it's identical to '선물명세'!A{row} for every one of the
   5 real rows -- every other column in this row is a live reference to
   선물명세. Replaced with a formula reference for the same reason as
   everywhere else in this project: so a future period's counterparty list
   is picked up automatically instead of needing a manual literal update.

2. N/O/P (the three check columns) only exist for the first 4 of the 5
   real counterparty rows in the real file -- the 5th (키움증권, row 18)
   has no check formulas at all, silently excluded. Same "range not
   extended when row count changed" pattern as pyeongga_jusik's stale
   COUNTIF range. Fixed here by generating N/O/P for every real row,
   driven by `num_rows` rather than a hardcoded count.
"""
from __future__ import annotations

from openpyxl.styles import Border, PatternFill, Side

from core.seonmul import SHEET_NAME as SEONMUL_SHEET_NAME

SHEET_NAME = "평가액검증(선물)"
KRX_SHEET_NAME = "KRX정보시스템_코스피200"

FIRST_DATA_ROW = 14
SOURCE_FIRST_DATA_ROW = 3  # 선물명세's own first data row

_PASSTHROUGH_COLUMNS = ["A", "B", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
# D (표준코드) has no "ColumnN" group label in the real file -- just omitted,
# not merged with C's like I10:J10/O12:P12 are elsewhere.
_GROUP_HEADER_COLUMNS = ["A", "B", "C", "E", "F", "G", "H", "I", "J", "K", "L", "M"]

ACCOUNTING_FMT = r'_-* #,##0_-;\-* #,##0_-;_-* "-"_-;_-@_-'

_TEST_LABEL_FILL = PatternFill("solid", fgColor="FFFFFF00")
_TEST_GROUP_BORDER = Border(top=Side(style="medium"), bottom=Side(style="thin"))
_TEST_LABEL_BORDER = Border(top=Side(style="thin"), bottom=Side(style="medium"))

_COLUMN_LABELS = {
    "A": "\t매매처        ",
    "B": "매매처명      ",
    "C": "종목코드",
    "D": "표준코드      ",
    "E": "종목명        ",
    "F": "포지션        ",
    "G": "계약수        ",
    "H": "정산가        ",
    "I": "평가금액      ",
    "J": "취득가액      ",
    "K": "정산손익      ",
    "L": "미수입금      ",
    "M": "미지급금      ",
}


def build(wb, num_rows: int):
    ws = wb.create_sheet(SHEET_NAME)
    s = SEONMUL_SHEET_NAME

    ws["G1"] = "='C1_정산표'!B1"
    ws["G2"] = "='C1_정산표'!B2"
    ws["L2"] = "Prepared by : "
    ws["M2"] = "='C1_정산표'!H2"
    ws["G3"] = "평가액검증"
    ws["L3"] = "Reviewed by : "
    ws["M3"] = " "

    for i, col in enumerate(_GROUP_HEADER_COLUMNS, start=1):
        ws[f"{col}12"] = f"Column{i}"
    ws["N12"] = "Test 1"
    ws["O12"] = "Test 2"
    ws.merge_cells("O12:P12")
    for col in "NOP":
        ws[f"{col}12"].border = _TEST_GROUP_BORDER

    for col, label in _COLUMN_LABELS.items():
        ws[f"{col}13"] = label
    ws["N13"] = "평가금액정확성"
    ws["O13"] = "정산가검증 A"
    ws["P13"] = "정산가검증 B"
    for col in "NOP":
        ws[f"{col}13"].fill = _TEST_LABEL_FILL
        ws[f"{col}13"].border = _TEST_LABEL_BORDER

    for i in range(num_rows):
        row = FIRST_DATA_ROW + i
        src_row = SOURCE_FIRST_DATA_ROW + i
        for col in _PASSTHROUGH_COLUMNS:
            ws[f"{col}{row}"] = f"='{s}'!{col}{src_row}"
        ws[f"C{row}"] = f"=MID(D{row},4,8)"
        ws[f"N{row}"] = f"=G{row}*H{row}*250000=I{row}"
        ws[f"O{row}"] = (
            f"=INDEX({KRX_SHEET_NAME}!$A$1:$L$14,"
            f"MATCH('{SHEET_NAME}'!C{row},{KRX_SHEET_NAME}!$A:$A,0),"
            f"MATCH({KRX_SHEET_NAME}!$C$1,{KRX_SHEET_NAME}!$1:$1,0))"
        )
        ws[f"P{row}"] = f'=IF(H{row}=0,"TRUE",IF(H{row}=O{row},"TRUE","FALSE"))'

        for col in "GHIJKLM":
            ws[f"{col}{row}"].number_format = ACCOUNTING_FMT

    return ws

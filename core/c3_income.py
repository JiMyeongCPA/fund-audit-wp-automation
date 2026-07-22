"""C3_수익: "이자수익 재계산" -- an analytical review that estimates each
interest-income category's expected income (평균잔액 x 평균이자율) and
compares it to the amount actually recorded in 통합기준가격대장, concluding
"적정" if the total difference is under 5%, "부적정" otherwise.

Structurally this sheet is much smaller than it looks (max_row reports 470,
but everything past row 26 is empty -- leftover cell formatting, not real
content). Row 8 ("이자수익 ARP") is a dead label with nothing else around it
in the real file; copied verbatim, not something we build logic for.

The real file hardcodes literal 0 for 외화채권/콜/정기예금/REPO (rows
18,20,22,23). 일별자산내역 has a column for every one of these, and three of
them (외화채권/콜/정기예금) are genuinely, literally zero on all 365 days --
verified directly, not just averaging to zero -- so those get a real AVERAGE
formula instead of a hardcoded literal (같은 원칙 C2의 콜론/매입어음 블록에서
이미 적용: 지금은 0이어도 실제 잔액이 생기면 자동으로 반영되게).

REPO (row 23) is the exception: its 일별자산내역 column is non-zero on 6 of
365 days, spiking to 9B won -- but that pattern (a brief spike, not a
sustained balance) matches short-term repo *borrowing* for cash management,
the same kind of transaction as C2's 5B REPO매도 liability, not repo
*lending* that would earn this fund interest income. So it stays a literal
0 here, same as the original -- wiring it up would misclassify a liability
balance as interest-earning income. 어음 (row 21) has no matching column in
일별자산내역 at all, so it's a literal 0 for a different reason (nothing to
average).

외화예금 (row 25) gets the same upgrade: the real file leaves it as just a
label + recorded amount, no estimate columns at all (verified directly --
C/D/E/F/G/I/J are genuinely empty there, not a mistake on our end). But
일별자산내역 has a matching column (L/M), genuinely zero on all 365 days
like 외화채권/콜/정기예금, so it gets the full row treatment here instead of
staying incomplete.

차이비율 (J column) uses IFERROR(...,0) uniformly instead of the real
file's per-row choice between a live formula and a hardcoded 0 (avoiding a
0/0 error): same result today, but correct automatically if a category that's
zero now ever becomes non-zero.

Borders/fills/number formats below are all copied from the real file's own
styling (theme colors read directly off its cells, not guessed): header row
(theme 6), rows with a real balance (theme 5, 채권/예금/위탁증거금), the
차이/차이비율 columns which are always highlighted (theme 4) regardless of
whether the row itself has a balance, and the 결론 cell (theme 3).
"""
from __future__ import annotations

from openpyxl.styles import Border, Color, PatternFill, Side

SHEET_NAME = "C3_수익"
MASTER_SHEET_NAME = "통합기준가격대장(결산후)"
YUDONG_SHEET_NAME = "국내유동명세"
ILBYEOLJASAN_SHEET_NAME = "일별자산내역"

I = ILBYEOLJASAN_SHEET_NAME
Y = YUDONG_SHEET_NAME

# (row, master_row, C-column 평균잔액 formula or None if no matching PBC column, is_active)
_ROW_CONFIG = [
    (17, 344, f"=AVERAGE({I}!J2:J366)-AVERAGE({I}!K2:K366)", True),  # 채권
    (18, 345, f"=AVERAGE({I}!O2:O366)-AVERAGE({I}!P2:P366)", False),  # 외화채권
    (19, 346, f"=AVERAGE({I}!B2:B366)-AVERAGE({I}!C2:C366)", True),  # 예금
    (20, 347, f"=AVERAGE({I}!D2:D366)-AVERAGE({I}!E2:E366)", False),  # 콜
    (21, 348, None, False),  # 어음 -- 일별자산내역에 대응 컬럼 자체가 없음
    (22, 349, f"=AVERAGE({I}!F2:F366)-AVERAGE({I}!G2:G366)", False),  # 정기예금
    (23, 350, None, False),  # REPO -- 컬럼은 있지만 부채성 단기차입 스파이크라 이자수익 아님
    (24, 351, f"=SUM({Y}!K5:K8)", True),  # 위탁증거금
    (25, 352, f"=AVERAGE({I}!L2:L366)-AVERAGE({I}!M2:M366)", False),  # 외화예금
]

AMOUNT_FMT = r"#,##0_);[Red]\(#,##0\)"
PERCENT_FMT = "0.00%"

_HEADER_FILL = PatternFill("solid", fgColor=Color(theme=6, tint=0.7999816888943144))
_ACTIVE_FILL = PatternFill("solid", fgColor=Color(theme=5, tint=0.7999816888943144))
_COMPUTED_FILL = PatternFill("solid", fgColor=Color(theme=4, tint=0.7999816888943144))
_CONCLUSION_FILL = PatternFill("solid", fgColor=Color(theme=3, tint=0.5999938962981048))

_HEADER_BORDER = Border(top=Side(style="thin"), bottom=Side(style="double"))
_ROW_BOTTOM_BORDER = Border(bottom=Side(style="thin"))
_CONCLUSION_BORDER = Border(top=Side(style="medium"), bottom=Side(style="medium"))
_TOTAL_BORDER = Border(top=Side(style="thin"), bottom=Side(style="thin"))


def build(wb, rate_items, day_split: tuple[int, int]):
    """`rate_items` is the list returned by
    interest_rate_source.extract_interest_rates (9 items, same row order as
    the real file: 채권/외화채권/예금/콜/어음/정기예금/REPO/위탁증거금/외화예금).
    `day_split` is (beginning_days, ending_days) from extract_day_split.
    """
    ws = wb.create_sheet(SHEET_NAME)
    m = MASTER_SHEET_NAME
    begin_days, end_days = day_split

    ws["A1"] = "='C1_정산표'!B1"
    ws["A2"] = "='C1_정산표'!B2"
    ws["F2"] = "Prepared by : "
    ws["G2"] = "='C1_정산표'!H2"
    ws["F3"] = "Reviewed by : "
    ws["A3"] = "수익"

    ws["B8"] = "이자수익 ARP"

    ws["B13"] = "이자수익 재계산"
    ws["C13"] = "결론"
    ws["D13"] = '=IF(J26<0.05,"적정","부적정")'
    ws["C13"].fill = _CONCLUSION_FILL
    ws["D13"].fill = _CONCLUSION_FILL
    ws["D13"].border = _CONCLUSION_BORDER

    headers = ["이자수익구분", "평균잔액", "기초이자율", "기말이자율", "평균이자율", "추정이자수익", "정산표상 이자수익", "차이", "차이비율"]
    for col, header in zip("BCDEFGHIJ", headers):
        cell = ws[f"{col}16"]
        cell.value = header
        cell.fill = _HEADER_FILL
        cell.border = _HEADER_BORDER

    for i, (row, master_row, c_formula, is_active) in enumerate(_ROW_CONFIG):
        item = rate_items[i]
        ws[f"B{row}"] = f"='{m}'!F{master_row}"
        ws[f"C{row}"] = c_formula if c_formula is not None else 0
        ws[f"D{row}"] = item.beginning_rate
        ws[f"E{row}"] = item.ending_rate
        ws[f"F{row}"] = f"=D{row}*{begin_days}/365+E{row}*{end_days}/365"
        ws[f"G{row}"] = f"=C{row}*F{row}"
        ws[f"H{row}"] = f"='{m}'!G{master_row}"
        ws[f"I{row}"] = f"=G{row}-H{row}"
        ws[f"J{row}"] = f"=IFERROR(I{row}/G{row},0)"

        for col in "CGHI":
            ws[f"{col}{row}"].number_format = AMOUNT_FMT
        for col in "DEFJ":
            ws[f"{col}{row}"].number_format = PERCENT_FMT
        for col in "BCDEFGHIJ":
            ws[f"{col}{row}"].border = _ROW_BOTTOM_BORDER
        if is_active:
            for col in "BCDEFGHI":
                ws[f"{col}{row}"].fill = _ACTIVE_FILL
        ws[f"J{row}"].fill = _COMPUTED_FILL

    ws["B26"] = "합계"
    ws["G26"] = "=SUM(G17:G25)"
    ws["H26"] = "=SUM(H17:H25)"
    ws["I26"] = "=SUM(I17:I25)"
    ws["J26"] = "=IFERROR(I26/G26,0)"
    for col in "GHI":
        ws[f"{col}26"].number_format = AMOUNT_FMT
        ws[f"{col}26"].fill = _COMPUTED_FILL
        ws[f"{col}26"].border = _TOTAL_BORDER
    ws["J26"].number_format = PERCENT_FMT
    ws["J26"].fill = _COMPUTED_FILL
    ws["J26"].border = _TOTAL_BORDER

    return ws

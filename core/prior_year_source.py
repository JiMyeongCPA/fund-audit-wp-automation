"""Build a standalone "전기 재무제표" workbook out of the 전기(prior-year)
column already sitting in a reference workpaper's C1_정산표 -- a stand-in for
"read the real prior-year audit report" until that's actually automatable
(the PDF's text is unreadable by any library we tried; the real DSD source
isn't something we have access to).

This deliberately decouples two separate problems: *getting* 전기 재무제표
data (still an open problem) from *using* 전기 재무제표 data to populate this
year's C1_정산표 (something we can build and test right now against a
reconstructed file that has the right shape).

C1_정산표 layout notes this relies on:
- Detail line items keep their 전기 값 in column I.
- Section subtotals (Ⅰ./Ⅱ./Ⅲ./Ⅳ., 자산 총계 등) keep theirs in column J
  instead (their 당기 counterpart lives in H rather than G, for the same
  reason).
- Column E is "공시"/"공시제외"; only "공시" rows are ever actually shown in
  the real published financial statements, so those are the only ones worth
  reconstructing here.
"""
from __future__ import annotations

from dataclasses import dataclass

import openpyxl

BS_RANGE = (9, 66)  # 재무상태표
IS_RANGE = (73, 118)  # <공고용 IS> 손익계산서

LABEL_COL = 6  # F열 과목
DISCLOSURE_COL = 5  # E열 공시여부
DETAIL_VALUE_COL = 9  # I열
SECTION_VALUE_COL = 10  # J열


@dataclass
class FinancialLineItem:
    section: str  # "재무상태표" or "손익계산서"
    row: int  # row this was read from in the reference C1_정산표
    label: str
    amount: float


@dataclass
class CapitalStatementRow:
    label: str
    principal: float  # 원본 (G열)
    retained_earnings: float  # 이익잉여금 (H열)
    total: float  # 총계 (I열)


# Rows 126-132: the block C1_정산표 already carries as "전기" capital
# movements (그 해의 설정/해지/손익/조정 내역). The row-126/132 captions in
# the raw sheet are stale leftover dates (e.g. "2022년 1월 1일(전기초)") from
# however many cycles ago this template was last rolled forward without
# anyone updating the label text -- the *numbers* are correct (verified
# against the actual prior-year audit report elsewhere), so only those two
# captions get normalized to plain "전기초"/"전기말" here rather than carried
# over verbatim.
CAPITAL_STATEMENT_ROWS = list(range(126, 133))
CAPITAL_LABEL_COL = 6  # F열
CAPITAL_PRINCIPAL_COL = 7  # G열
CAPITAL_RETAINED_EARNINGS_COL = 8  # H열
CAPITAL_TOTAL_COL = 9  # I열


def extract_prior_year_capital_statement(reference_wb) -> list[CapitalStatementRow]:
    ws = reference_wb["C1_정산표"]
    rows = []
    for i, row in enumerate(CAPITAL_STATEMENT_ROWS):
        label = ws.cell(row, CAPITAL_LABEL_COL).value
        if i == 0:
            label = "전기초"
        elif i == len(CAPITAL_STATEMENT_ROWS) - 1:
            label = "전기말"
        rows.append(
            CapitalStatementRow(
                label=label,
                principal=ws.cell(row, CAPITAL_PRINCIPAL_COL).value or 0,
                retained_earnings=ws.cell(row, CAPITAL_RETAINED_EARNINGS_COL).value or 0,
                total=ws.cell(row, CAPITAL_TOTAL_COL).value or 0,
            )
        )
    return rows


def extract_prior_year_financials(reference_wb) -> list[FinancialLineItem]:
    ws = reference_wb["C1_정산표"]
    items: list[FinancialLineItem] = []

    for section_name, (first_row, last_row) in [
        ("재무상태표", BS_RANGE),
        ("손익계산서", IS_RANGE),
    ]:
        for row in range(first_row, last_row + 1):
            label = ws.cell(row, LABEL_COL).value
            if not isinstance(label, str):
                continue
            if ws.cell(row, DISCLOSURE_COL).value != "공시":
                continue
            value = ws.cell(row, DETAIL_VALUE_COL).value
            if value is None:
                value = ws.cell(row, SECTION_VALUE_COL).value
            if not isinstance(value, (int, float)):
                continue
            items.append(FinancialLineItem(section_name, row, label, value))

    return items


def write_prior_year_workbook(
    items: list[FinancialLineItem],
    output_path,
    capital_rows: list[CapitalStatementRow] | None = None,
) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for section_name in ("재무상태표", "손익계산서"):
        ws = wb.create_sheet(section_name)
        ws.append(["과목", "금액"])
        for item in items:
            if item.section == section_name:
                ws.append([item.label, item.amount])
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20

    if capital_rows:
        ws = wb.create_sheet("자본변동표")
        ws.append(["구분", "원본", "이익잉여금", "총계"])
        for row in capital_rows:
            ws.append([row.label, row.principal, row.retained_earnings, row.total])
        ws.column_dimensions["A"].width = 20
        for col in "BCD":
            ws.column_dimensions[col].width = 20

    wb.save(str(output_path))

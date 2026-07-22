"""C5_F.N > "1. 원본": 주석 disclosure of the trust's 원본(principal/units
outstanding) -- 당기말/전기말 발행좌수 and 원본 총액, narrated as two
TEXT()-embedded sentences (matching the real file's own approach: the
numbers are computed once in G/H columns, then quoted into the sentence via
TEXT() rather than typed twice).

Both numbers trace back to cells already fully wired elsewhere in this
workbook, no new PBC source needed:

- 당기 좌수/원본 (G146 in 'C1_정산표') -- the 당기말 총계 of the CURRENT-year
  capital-movement schedule (rows 136-146), already copied verbatim.
- 전기 좌수/원본 (G136 in 'C1_정산표') -- that same schedule's own 당기초
  row, which is itself just `=G132` (전기말 원본 = 당기초 원본, same
  balance) pointing at the PRIOR-year schedule (rows 126-132) whose detail
  rows are the ones prior_year_source reconstructs. So "전기" here doesn't
  need its own separate reconstruction -- it rides on the reconstruction C1
  already has.

Scope note: the real file's row 1023 onward is a "종류별 발행좌수" (by unit
class) breakdown table that pulls from a separate class-level dividend
schedule elsewhere in this sheet (rows 1058-1090). That whole mechanism
assumes a multi-class fund with its own per-class 기준가격대장 (there's a
literal comment in the real file at N1060: ">>> 각 클래스별
기준가격대장(결산후)에서 추출") -- 샘플펀드 only has one unit class, so
there's no such per-class ledger to extract from. The real file's own
class-breakdown table reflects this: several #DIV/0! errors and a
"정산표대사" reconciliation check that's cached as FALSE. Not a PBC gap to
fill later (like the deferred sheets in project memory) -- structurally
inapplicable to this fund, same category as C2's 외화위탁증거금. The
narrative sentence's closing clause ("종류별 발행좌수는 다음과 같습니다")
is kept verbatim from the real disclosure text rather than edited. The
table shape itself (header + a few blank rows + 합계) is still built, since
a form with nothing to fill in is still a form -- just no formulas, and no
정산표대사 check row (that check is meaningless with no real class data,
unlike c5_ibunbae_note's, which is retained because that one does have a
real reconciliation target).
"""
from __future__ import annotations

from openpyxl.styles import Color, PatternFill

from core.c5_styles import box_border, left_rule, style_header_row, style_total_row
from core.c5_styles import BOLD as _BOLD

SETTLEMENT_SHEET_NAME = "C1_정산표"

CURRENT_PERIOD_TOTAL_ROW = 146  # 'C1_정산표'!G146: 당기말 원본
PRIOR_PERIOD_TOTAL_ROW = 136  # 'C1_정산표'!G136: 전기말(=당기초) 원본
ISSUE_PRICE_PER_UNIT = 10000  # 1좌당 발행가액, 리터럴 (신탁약관 고정값)

AMOUNT_FMT = r"#,##0\ ;\(#,##0\);\-\ ;@"

_UNIT_LABEL_FILL = PatternFill("solid", fgColor=Color(theme=7, tint=0.7999816888943144))
_PRICE_LABEL_FILL = PatternFill("solid", fgColor="FFFFC000")
_SENTENCE_FILL = PatternFill("solid", fgColor="FFFFFF00")
_CLASS_BREAKDOWN_BUFFER_ROWS = 3


def build(ws, start_row: int = 1020) -> int:
    """Writes the "1. 원본" note (narrative + total only, no 종류별
    breakdown -- see module docstring) starting at `start_row`. Returns the
    next free row after this section."""
    s = SETTLEMENT_SHEET_NAME
    r = start_row

    ws.cell(r, 3, "1. 원본").font = _BOLD
    ws.cell(r, 7, "원단위").fill = _UNIT_LABEL_FILL
    ws.cell(r, 8, "천원단위").fill = _UNIT_LABEL_FILL
    ws.cell(r, 9, "좌당").fill = _PRICE_LABEL_FILL
    left_rule(ws, r)

    current_row = r + 1
    prior_row = current_row + 1

    ws.cell(current_row, 3, (
        f'="당기말 및 전기말 현재 투자신탁의 발행좌수는 각각 "&TEXT(G{current_row},"#,###")'
        f'&"좌 및 "&TEXT(G{prior_row},"#,###")&"좌이며, "'
    ))
    ws.cell(current_row, 6, "당기")
    ws.cell(current_row, 7, f"='{s}'!G{CURRENT_PERIOD_TOTAL_ROW}/I{current_row}")
    ws.cell(current_row, 8, f"=ROUND(G{current_row}/1000,0)")
    ws.cell(current_row, 9, ISSUE_PRICE_PER_UNIT)

    ws.cell(prior_row, 3, (
        f'="1좌당 발행가액은 "&TEXT(I{current_row},"#,###")&"원으로 원본은 각각 "'
        f'&TEXT(H{current_row}*I{current_row},"#,###")&"천원 및 "'
        f'&TEXT(H{prior_row}*I{prior_row},"#,###")&"천원이며 종류별 발행좌수는 다음과 같습니다."'
    ))
    ws.cell(prior_row, 6, "전기")
    ws.cell(prior_row, 7, f"='{s}'!G{PRIOR_PERIOD_TOTAL_ROW}/I{prior_row}")
    ws.cell(prior_row, 8, f"=ROUND(G{prior_row}/1000,0)")
    ws.cell(prior_row, 9, ISSUE_PRICE_PER_UNIT)

    for row in (current_row, prior_row):
        for col in (3, 4, 5):
            ws.cell(row, col).fill = _SENTENCE_FILL
        for col in (3, 4, 5, 6, 7, 8, 9):
            ws.cell(row, col).border = box_border(col == 3)
        for col in (7, 8, 9):
            ws.cell(row, col).number_format = AMOUNT_FMT
        left_rule(ws, row)

    class_header_row = prior_row + 2
    for col, label in zip("CDE", ["종류", "당기", "전기"]):
        ws.cell(class_header_row, ord(col) - ord("A") + 1, label)
    style_header_row(ws, class_header_row, 3, 5)
    left_rule(ws, prior_row + 1)

    class_total_row = class_header_row + 1 + _CLASS_BREAKDOWN_BUFFER_ROWS
    ws.cell(class_total_row, 3, "합계")
    style_total_row(ws, class_total_row, 3, 5)
    for row in range(class_header_row + 1, class_total_row):
        left_rule(ws, row)
    left_rule(ws, class_total_row)

    return class_total_row + 2

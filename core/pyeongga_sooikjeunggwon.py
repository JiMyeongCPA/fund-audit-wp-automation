"""평가액검증(수익증권): row-by-row valuation check for every domestic
collective-investment-security (ETF) position, same idea as
pyeongga_jusik.py but pulling a different, non-sequential set of columns
from '수익증권명세' (종목코드/표준코드/종목명/좌수/취득가액/기준가격/평가금액/
평가손익 -- columns A,B,C,G,K,I,J,L of the source, not simply its first 8)
and checking against 'KRX정보시스템_ETF' instead of KRS정보시스템_주식.

- 평가금액정확성 (I): 좌수 x 기준가격 == 평가금액.
- 정산가검증 A (J): looks up 기준가격 in 'KRX정보시스템_ETF' via INDEX/MATCH
  on 종목명 (same self-referencing C1 trick as pyeongga_jusik's E1: C1 holds
  the literal '종가', matching KRX정보시스템_ETF's own C-column header).
- 정산가검증 B (K): TRUE/FALSE comparing F (수익증권명세's 기준가격) against
  J -- guarded the same way as 종가검증 B: if F is 0 (no live position),
  skip straight to "TRUE" instead of comparing.

This fund holds exactly one ETF (종목ETF001) this period, but the row count
is still driven by `num_rows` (len of the parsed 수익증권명세 df) rather
than hardcoded, matching the row-count-driven approach used everywhere else
in this project.
"""
from __future__ import annotations

from openpyxl.styles import Border, Color, PatternFill, Side

from core.sooikjeunggwon import SHEET_NAME as SOOIKJEUNGGWON_SHEET_NAME

SHEET_NAME = "평가액검증(수익증권)"
KRX_SHEET_NAME = "KRX정보시스템_ETF"

FIRST_DATA_ROW = 7
SOURCE_FIRST_DATA_ROW = 3  # 수익증권명세's own first data row

# (own column letter, source column letter) for A-H -- non-sequential:
# 좌수/취득가액/기준가격/평가금액/평가손익 sit at G/K/I/J/L in the source.
_COLUMN_MAP = list(zip("ABCDEFGH", "ABCGKIJL"))

ACCOUNTING_FMT = r'_-* #,##0_-;\-* #,##0_-;_-* "-"_-;_-@_-'
ACCOUNTING_FMT_2DP = r'_-* #,##0.00_-;\-* #,##0.00_-;_-* "-"_-;_-@_-'

_TEST_LABEL_FILL = PatternFill("solid", fgColor="FFFFFF00")
_TEST_GROUP_BORDER = Border(bottom=Side(style="medium"))
_TEST_LABEL_BORDER = Border(top=Side(style="thin"), bottom=Side(style="medium"))
_ACTIVE_FILL = PatternFill("solid", fgColor=Color(theme=5, tint=0.7999816888943144))

_COLUMN_LABELS = [
    "종목코드         ",
    "표준코드         ",
    "종목명           ",
    "좌수             ",
    "취득가액         ",
    "기준가격         ",
    "평가금액         ",
    "평가손익         ",
]


def build(wb, num_rows: int):
    ws = wb.create_sheet(SHEET_NAME)
    s = SOOIKJEUNGGWON_SHEET_NAME

    ws["B1"] = "='C1_정산표'!B1"
    ws["B2"] = "='C1_정산표'!B2"
    ws["G2"] = "Prepared by : "
    ws["H2"] = "='C1_정산표'!H2"
    ws["B3"] = "평가액검증"
    ws["G3"] = "Reviewed by : "
    ws["H3"] = " "

    for col, label in zip("ABCDEFGH", [f"Column{i}" for i in range(1, 9)]):
        ws[f"{col}5"] = label
    for col in "IJK":
        ws[f"{col}5"].border = _TEST_GROUP_BORDER

    for col, label in zip("ABCDEFGH", _COLUMN_LABELS):
        ws[f"{col}6"] = label
    ws["I6"] = "평가금액정확성"
    ws["J6"] = "정산가검증 A"
    ws["K6"] = "정산가검증 B"
    for col in "IJK":
        ws[f"{col}6"].fill = _TEST_LABEL_FILL
        ws[f"{col}6"].border = _TEST_LABEL_BORDER

    for i in range(num_rows):
        row = FIRST_DATA_ROW + i
        src_row = SOURCE_FIRST_DATA_ROW + i
        for col, src_col in _COLUMN_MAP:
            ws[f"{col}{row}"] = f"='{s}'!{src_col}{src_row}"
        ws[f"I{row}"] = f"=D{row}*F{row}=G{row}"
        ws[f"J{row}"] = (
            f"=INDEX({KRX_SHEET_NAME}!$A$1:$R$2000,"
            f"MATCH('{SHEET_NAME}'!C{row},{KRX_SHEET_NAME}!$B:$B,0),"
            f"MATCH({KRX_SHEET_NAME}!$C$1,{KRX_SHEET_NAME}!$1:$1,0))"
        )
        ws[f"K{row}"] = f'=IF(F{row}=0,"TRUE",IF(F{row}=J{row},"TRUE","FALSE"))'

        ws[f"I{row}"].number_format = ACCOUNTING_FMT
        ws[f"J{row}"].number_format = ACCOUNTING_FMT_2DP
        ws[f"K{row}"].number_format = ACCOUNTING_FMT
        ws[f"I{row}"].fill = _ACTIVE_FILL
        ws[f"K{row}"].fill = _ACTIVE_FILL

    return ws

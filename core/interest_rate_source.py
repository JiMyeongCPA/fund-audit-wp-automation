"""Build a standalone "이자율 기초자료" workbook out of C3_수익's own
기초/기말 이자율 (D/E columns, rows 17-25) and the day-count split used to
average them (F column formula, e.g. `=D17*100/365+E17*265/365`) -- a
stand-in for "look up the real benchmark rate and the date it changed"
until that's actually automatable, the same way prior_year_source.py stands
in for "read the real prior-year audit report."

This decouples "where do the interest rates come from" (still an open
problem -- no PBC source for a benchmark rate) from "how do we use interest
rates to build C3_수익" (buildable and testable right now against a
reconstructed file with the right shape).

C3_수익 layout notes this relies on:
- Rows 17-25 are the 9 fixed interest-income categories (채권/외화채권/예금/
  콜/어음/정기예금/REPO/위탁증거금/외화예금); B column labels come from
  '통합기준가격대장(결산후)'!F344:F352, not read here.
- D/E (기초/기말이자율) are mostly literal numbers a preparer typed in from
  an external rate source; one row (예금, row 19 here) instead has E as a
  live formula pulling from 국내유동명세 -- data_only=True resolves that to
  its current numeric value like any other cell.
- The day-count split (e.g. 100/265) is embedded inside the F-column
  formula text itself, not a separate cell -- pulled out with a regex
  rather than read as a value.
"""
from __future__ import annotations

import re
from dataclasses import dataclass

import openpyxl

RATE_ROWS = list(range(17, 26))
LABEL_COL = 2  # B열
BEGIN_RATE_COL = 4  # D열 기초이자율
END_RATE_COL = 5  # E열 기말이자율
DAY_SPLIT_FORMULA_COL = 6  # F열, e.g. =D17*100/365+E17*265/365

_DAY_SPLIT_PATTERN = re.compile(r"\*(\d+)/365\+[A-Z]+\d+\*(\d+)/365")


@dataclass
class InterestRateItem:
    label: str
    beginning_rate: float
    ending_rate: float


def extract_interest_rates(reference_wb_values) -> list[InterestRateItem]:
    """`reference_wb_values` must be data_only=True so a rate that's a live
    formula (like 예금's E) resolves to its current number instead of
    formula text."""
    ws = reference_wb_values["C3_수익"]
    items = []
    for row in RATE_ROWS:
        label = ws.cell(row, LABEL_COL).value
        if not isinstance(label, str):
            continue
        items.append(
            InterestRateItem(
                label=label,
                beginning_rate=ws.cell(row, BEGIN_RATE_COL).value or 0,
                ending_rate=ws.cell(row, END_RATE_COL).value or 0,
            )
        )
    return items


def extract_day_split(reference_wb_formulas) -> tuple[int, int]:
    """`reference_wb_formulas` must be data_only=False -- the day counts are
    literal numbers embedded in the F-column formula text, not their own
    cells. Returns (beginning_days, ending_days); scans rows 17-25 for the
    first row that actually has this formula (rows with zero balance just
    have a literal 0 in F instead)."""
    ws = reference_wb_formulas["C3_수익"]
    for row in RATE_ROWS:
        formula = ws.cell(row, DAY_SPLIT_FORMULA_COL).value
        if not isinstance(formula, str):
            continue
        match = _DAY_SPLIT_PATTERN.search(formula)
        if match:
            return int(match.group(1)), int(match.group(2))
    raise ValueError("no row in C3_수익 has the expected day-split formula")


def write_interest_rate_workbook(
    items: list[InterestRateItem], day_split: tuple[int, int], output_path
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "이자율"

    ws.append(["구분", "기초이자율", "기말이자율"])
    for item in items:
        ws.append([item.label, item.beginning_rate, item.ending_rate])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15

    notes_ws = wb.create_sheet("적용일수")
    notes_ws.append(["기초금리 적용일수", "기말금리 적용일수"])
    notes_ws.append(list(day_split))

    wb.save(str(output_path))

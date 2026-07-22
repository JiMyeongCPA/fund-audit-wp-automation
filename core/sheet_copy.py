"""Copy one worksheet's content into a different workbook.

openpyxl can duplicate a sheet within the *same* workbook (`Workbook.copy_worksheet`)
but has no built-in way to carry a sheet across two separate Workbook objects,
which is what we need to pull '통합기준가격대장(결산후)' out of a reference
workpaper (전기 조서) and into the fresh workbook we're assembling.
"""
from __future__ import annotations

from copy import copy

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def copy_sheet(
    src_wb: Workbook,
    sheet_name: str,
    dst_wb: Workbook,
    new_name: str | None = None,
    unhide: bool = False,
) -> Worksheet:
    """Copy `sheet_name` from `src_wb` into `dst_wb` as a new sheet (named
    `new_name`, or `sheet_name` if omitted). Carries over cell values/formulas,
    per-cell styling, column widths, row heights, hidden rows/columns, merged
    ranges, and the sheet's own hidden/visible state.

    `unhide=True` drops all of that hidden-ness instead of carrying it over --
    every row/column stays visible and the sheet itself is forced visible,
    regardless of what the source had hidden.
    """
    src_ws = src_wb[sheet_name]
    dst_ws = dst_wb.create_sheet(new_name or sheet_name)

    for row in src_ws.iter_rows():
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dst_cell.font = copy(cell.font)
                dst_cell.fill = copy(cell.fill)
                dst_cell.border = copy(cell.border)
                dst_cell.alignment = copy(cell.alignment)
                dst_cell.protection = copy(cell.protection)
                dst_cell.number_format = cell.number_format

    for col_letter, dim in src_ws.column_dimensions.items():
        dst_dim = dst_ws.column_dimensions[col_letter]
        dst_dim.width = dim.width
        dst_dim.hidden = False if unhide else dim.hidden

    for row_idx, dim in src_ws.row_dimensions.items():
        dst_dim = dst_ws.row_dimensions[row_idx]
        dst_dim.height = dim.height
        dst_dim.hidden = False if unhide else dim.hidden

    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    dst_ws.sheet_state = "visible" if unhide else src_ws.sheet_state
    return dst_ws


def copy_row_range(
    src_ws: Worksheet,
    dst_ws: Worksheet,
    first_row: int,
    last_row: int,
) -> None:
    """Copy just rows `first_row`..`last_row` (values/formulas + per-cell
    style) from `src_ws` into `dst_ws` at the same row/column coordinates --
    unlike `copy_sheet`, this writes into a sheet that already exists (e.g.
    C2_자산부채평가's 요약표 rows, layered on top of detail blocks built
    separately), rather than creating a new one.
    """
    for row in src_ws.iter_rows(min_row=first_row, max_row=last_row):
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dst_cell.font = copy(cell.font)
                dst_cell.fill = copy(cell.fill)
                dst_cell.border = copy(cell.border)
                dst_cell.alignment = copy(cell.alignment)
                dst_cell.protection = copy(cell.protection)
                dst_cell.number_format = cell.number_format

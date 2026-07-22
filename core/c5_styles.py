"""Shared cell styling for C5_F.N note modules, read directly off the real
file's own cells (theme colors, border weights) rather than guessed --
same theme values already established in c3_income.py/c4_fee.py/
c5_bosu_note.py: header rows use theme 8, totals use theme 4, the tie-out/
check row uses theme 0, and every data row gets a thin grid box with a
medium rule down column C for the whole section.
"""
from __future__ import annotations

from openpyxl.styles import Border, Color, Font, PatternFill, Side

BOLD = Font(bold=True)

# Same convention as c2_blocks.BUFFER_ROWS: every PBC-row-count-driven
# listing (지분증권/채무증권/수익증권/파생상품/매도유가증권 등) gets N real
# rows + 3 blank buffer rows before its 합계, so next period's row count
# growing or shrinking doesn't require rebuilding the block by hand.
BUFFER_ROWS = 3

HEADER_FILL = PatternFill("solid", fgColor=Color(theme=8, tint=-0.249977111117893))
TOTAL_FILL = PatternFill("solid", fgColor=Color(theme=4, tint=0.7999816888943144))
CHECK_FILL = PatternFill("solid", fgColor=Color(theme=0, tint=-0.0499893185216834))
SUBTOTAL_FILL = PatternFill("solid", fgColor=Color(theme=7, tint=0.7999816888943144))

THIN = Side(style="thin")
MEDIUM = Side(style="medium")
HEADER_BORDER = Border(bottom=THIN)


def box_border(is_left_edge: bool = False) -> Border:
    return Border(top=THIN, bottom=THIN, right=THIN, left=MEDIUM if is_left_edge else THIN)


def style_header_row(ws, row: int, first_col: int, last_col: int) -> None:
    for col in range(first_col, last_col + 1):
        cell = ws.cell(row, col)
        cell.fill = HEADER_FILL
        cell.border = box_border(col == first_col)


def style_data_row(ws, row: int, first_col: int, last_col: int) -> None:
    for col in range(first_col, last_col + 1):
        ws.cell(row, col).border = box_border(col == first_col)


def style_total_row(ws, row: int, first_col: int, last_col: int) -> None:
    for col in range(first_col, last_col + 1):
        cell = ws.cell(row, col)
        cell.fill = TOTAL_FILL
        cell.border = box_border(col == first_col)


def style_subtotal_row(ws, row: int, first_col: int, last_col: int) -> None:
    for col in range(first_col, last_col + 1):
        cell = ws.cell(row, col)
        cell.fill = SUBTOTAL_FILL
        cell.border = box_border(col == first_col)


def left_rule(ws, row: int, col: int = 3) -> None:
    """A medium vertical rule down `col` (column C by default) for a single
    row -- call across every row of a section to run it the full height."""
    cell = ws.cell(row, col)
    cell.border = Border(left=MEDIUM, top=cell.border.top, bottom=cell.border.bottom, right=cell.border.right)


def run_left_rule(ws, first_row: int, last_row: int, col: int = 3) -> None:
    """Apply `left_rule` to every row in [first_row, last_row] -- call this
    LAST, after all other styling, so no gap row is ever missed (the bug a
    scattered, hand-picked set of left_rule calls kept causing)."""
    for row in range(first_row, last_row + 1):
        left_rule(ws, row, col)

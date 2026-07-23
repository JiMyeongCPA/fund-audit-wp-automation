"""1~3단계 전체 워크플로우 백엔드. FastAPI + React 프론트엔드(frontend/)에서
호출한다.

  - POST /api/build가 PBC 원본자료(11개 CSV + 전기 조서)를 업로드받아
    build_workpaper()를 실제로 돌리고, 그 결과(sample_output/조립결과.xlsx)를
    2단계 화면이 그대로 읽는다. 업로드를 생략하면 sample_data/의 표본으로
    대신 돌아간다.
  - 신규계정/전기값-드리프트 플래그는 build_workpaper()가 실제로 반환한
    값을 쓴다.
  - 승인형(특수관계자거래/담보제공자산) 항목의 근거검색은
    core.fee_contract_search.search()를 호출한다 (Gemini 임베딩) -- 사용량
    제한 등으로 실시간 AI를 못 쓰면 키워드 매칭 + 결정론적 폴백 제안으로
    대체하고, 사람이 예(승인)를 눌러야 그 값이 주석 시트에 반영된다.
"""
from __future__ import annotations

import io
import os
import shutil
import tempfile
from pathlib import Path

from dotenv import load_dotenv

load_dotenv()  # .env(gitignored)의 GEMINI_API_KEY -- 있으면 AI 근거검색이
# core.fee_contract_search.search()(진짜 임베딩)를 쓰고, 없으면 키워드매칭
# 폴백으로 조용히 내려간다(_search_candidates 참고).

import openpyxl
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from core.sheet_copy import copy_sheet

from core import pbc_library, prior_year_source
from core.assemble import build_workpaper
from core.flagged_items import FlaggedItem, build_flagged_items
from core.fee_contract_search import Candidate, Proposal, generate_proposal, read_document_pages, read_document_text

PROJECT_ROOT = Path(__file__).resolve().parent
FIXTURES_DIR = PROJECT_ROOT / "tests" / "fixtures"
TEUKSU_MANIFEST = FIXTURES_DIR / "synthetic_pbc" / "teuksu_demo.json"
DAMBO_MANIFEST = FIXTURES_DIR / "synthetic_pbc" / "dambo_demo.json"
# 업로드를 생략했을 때 쓰는 표본 입력은 sample_data/에 들어 있다(커밋됨) --
# 클론한 사람도 별도 자료 없이 바로 전체 흐름을 돌려볼 수 있다.
SAMPLE_DATA_DIR = PROJECT_ROOT / "sample_data"
BUILD_OUTPUT_PATH = PROJECT_ROOT / "sample_output" / "조립결과.xlsx"

# 승인형 항목의 근거검색·제안을 실시간 Gemini API로 돌릴지 여부. 기본은 꺼둔다
# (False) -- 무료/제한된 키에서 사용량 제한에 걸리면 1단계 실행(=/api/build)이
# 항목당 검색+제안+재시도 대기로 몇 분씩 멈추기 때문. 꺼진 상태에서는 검색은
# 키워드 매칭 폴백, 제안은 사전 정의된 결정론적 폴백을 써서 즉시·안정적으로
# 동작한다(README의 "AI API 사용량 제한에 대하여" 참고). 할당량이 넉넉한 키가
# 있으면 USE_LIVE_AI=1 로 켜서 실제 임베딩 검색+제안 생성을 쓴다.
USE_LIVE_AI = os.environ.get("USE_LIVE_AI") == "1"

# build_workpaper()의 각 인자 이름 -> sample_data/의 기본 파일. 업로드에서
# 해당 필드를 생략하면 이 파일로 대신 돌아간다 -- "일단 샘플로 전체 흐름
# 확인해보고, 나중에 진짜 파일로 교체" 두 가지를 같은 엔드포인트 하나로 처리.
SAMPLE_BUILD_INPUTS = {
    "gijunkagyeok": "기준가격대장(결산후).csv",
    "gungnaeyudong": "국내유동명세.csv",
    "gungnaejusik": "국내주식명세.csv",
    "pundbyeolmyeongse": "펀드별명세.csv",
    "sooikjeunggwon": "국내집합투자증권명세.csv",
    "seonmul": "국내선물명세.csv",
    "chaegwon": "채권명세.csv",
    "gajungpyeonggyunjwasu": "일별좌수순자산현황.csv",
    "bosubunbae": "판매보수내역.csv",
    "ilbyeoljasan": "일별자산내역.csv",
    "seoljeongheji": "설정해지내역.csv",
    "reference_workpaper": "전기조서.xlsx",
}

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def _stub_search(query: str, doc_paths: list[str], top_k: int = 3) -> list[Candidate]:
    keywords = [w for w in query.replace("(", " ").replace(")", " ").split() if len(w) > 1]
    scored = []
    for p in doc_paths:
        text = read_document_text(p)
        score = float(sum(text.count(k) for k in keywords))
        scored.append(Candidate(doc_path=p, text=text, score=score))
    scored.sort(key=lambda c: -c.score)
    return scored[:top_k]


def _search_candidates(item: FlaggedItem, use_ai: bool = True) -> tuple[list[Candidate], bool]:
    doc_paths = item.evidence_ref["doc_paths"]
    query = item.evidence_ref["query"]
    if use_ai and (os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")):
        try:
            from core.fee_contract_search import search

            return search(query, doc_paths, top_k=3), True
        except Exception:  # noqa: BLE001
            return _stub_search(query, doc_paths), False
    return _stub_search(query, doc_paths), False


# 채권/수익증권도 한때 여기서 fixture CSV를 직접 읽어 "결정형" 항목으로
# 만들었지만(_ORIGIN_SPECS), C2 데이터 셀들이 채권명세/수익증권명세를 수식으로
# 직접 참조한다는 게 거래소주식과 똑같이 확인되면서(2026-07-22) 제거했다 --
# 이제 예금/거래소주식과 동일하게 순수 수식 기반(/api/pbc-preview)으로만
# 처리된다. 백엔드에 "결정형" 항목으로 하드코딩해둘 이유가 없어졌다.


def _flagged_item_to_json(item: FlaggedItem, use_ai: bool = True) -> dict:
    d = {
        "kind": "flag",
        "id": item.id,
        "accountName": item.account_name,
        "reason": item.reason,
        "evidenceType": item.evidence_type,
        "decisionRequired": item.decision_required,
        "status": item.status,
        "isDemo": item.id.startswith("demo::"),
    }
    if item.evidence_type == "document":
        candidates, used_real_ai = _search_candidates(item, use_ai=use_ai)
        d["usedRealAi"] = used_real_ai
        d["candidates"] = [
            {
                "docName": Path(c.doc_path).name,
                "score": c.score,
                "text": c.text,
                "pageCount": len(read_document_pages(c.doc_path)) if c.doc_path.lower().endswith(".pdf") else None,
            }
            for c in candidates
        ]
        proposal, is_fallback = _build_proposal(item, candidates, used_real_ai)
        d["aiProposal"] = proposal.text if proposal else None
        d["aiProposalPage"] = proposal.page if proposal else None
        # 실제 Gemini 호출이 아니라(사용량 제한 등) 사전 정의된 폴백 제안을
        # 쓰는 중인지 -- 프론트가 작은 안내로 표시한다(README에도 명시).
        d["aiProposalIsFallback"] = is_fallback
        d["aiProposalUnavailable"] = proposal is None and len(candidates) > 0
    return d


# Gemini 사용량 제한 등으로 실제 LLM 제안을 못 만들 때 쓰는 결정론적 폴백.
# 원문(합성 계약서)에 실제로 있는 내용만 담아, AI가 살아있을 때 만드는 제안과
# 같은 취지가 되도록 손으로 작성했다. README에 "사용량 제한 시 이 폴백으로
# 대체된다(AI 기능 자체는 구현되어 있음)"라고 명시.
_DEMO_FALLBACK_PROPOSALS: dict[str, Proposal] = {
    "demo::특수관계자거래": Proposal(
        text=(
            "이사회 승인확인서에서 당기·전기 특수관계자거래 보수 및 미지급금 내역이 "
            "확인되어, 이를 '특수관계자와의 거래' 주석에 반영하는 것을 제안합니다."
        ),
        page=3,
    ),
    "demo::담보제공자산": Proposal(
        text=(
            "담보제공 약정서에서 종목002·종목003이 파생상품 위탁증거금 담보로 제공된 "
            "것이 확인되어, 담보로 제공된 종목·수량·장부가액을 '담보제공자산' 주석에 "
            "반영하는 것을 제안합니다."
        ),
        page=3,
    ),
}


def _build_proposal(
    item: FlaggedItem, candidates: list[Candidate], used_real_ai: bool
) -> tuple[Proposal | None, bool]:
    """(제안, 폴백여부)를 돌려준다. 실제 AI가 쓸 수 있으면 원문 근거 기반
    제안을 만들고, 사용량 제한 등으로 실패하면 결정론적 폴백으로 대체한다 --
    데모 항목(특수관계자거래/담보제공자산)은 어느 쪽이든 예/아니오로 반영할 수
    있어야 하므로, 폴백까지 두어 버튼이 항상 뜨게 한다."""
    if not candidates:
        return None, False
    if used_real_ai:
        try:
            proposal = generate_proposal(item.evidence_ref["query"], candidates[0].doc_path)
            if proposal.text:
                return proposal, False
        except Exception:  # noqa: BLE001
            pass
    fallback = _DEMO_FALLBACK_PROPOSALS.get(item.id)
    if fallback is not None:
        return fallback, True
    return None, False


# 거래소주식은 한때 여기서 AI 검색으로 데모했었지만(2026-07-22), C2 셀들이
# 국내주식명세를 수식으로 직접 참조한다는 게 확인되면서 더 정확하고 무료인
# 수식 기반 조회(core.pbc_library.find_by_sheet_name, /api/pbc-preview)로
# 대체됐다 -- AI 검색은 수식으로 연결 안 되는 항목(수수료비용 등)에만 남는다.

# ---- 검토 항목 스토어. 서버 시작 시엔 손으로 구성한 예시 신규계정으로 채우고,
# /api/build로 실제 조립을 한 번이라도 돌리면 그 진짜 build_result로 다시
# 채운다 -- 이후 승인/반려는 메모리에만 반영(데모 목적). ---------------------
_ITEMS_STORE: dict[str, dict] = {}

_FAKE_BUILD_RESULT = {
    "missing_account_codes": [{"코드": "5100999", "계정명": "파생상품평가이익(신규)"}],
    "drifted_prior_year_rows": [],
}


def _init_store(build_result: dict, use_ai: bool = True) -> None:
    """`use_ai=False`는 서버 부팅 시(아래 모듈 임포트 시점 호출)에만 쓴다 --
    AI 근거검색 항목(특수관계자거래/담보제공자산)마다 부팅할 때 실제 Gemini
    호출(검색+제안 생성, 요청량 제한에 걸리면 재시도 대기까지)을 동기로
    기다리면 서버가 뜨는 데만 몇 분씩 걸리거나 멈춘 것처럼 보이는 문제가
    있었다. 부팅 시 초기 상태는 어차피 실제 /api/build 한 번이면 다시 채워지는
    임시 표시일 뿐이라, 빠른 키워드매칭 폴백으로 채우고 실제 AI 호출은
    /api/build가 호출할 때만 쓴다."""
    _ITEMS_STORE.clear()
    teuksu_manifest = TEUKSU_MANIFEST if TEUKSU_MANIFEST.exists() else None
    dambo_manifest = DAMBO_MANIFEST if DAMBO_MANIFEST.exists() else None
    for flagged in build_flagged_items(
        build_result,
        teuksu_manifest_path=teuksu_manifest,
        dambo_manifest_path=dambo_manifest,
    ):
        _ITEMS_STORE[flagged.id] = _flagged_item_to_json(flagged, use_ai=use_ai)


_init_store(_FAKE_BUILD_RESULT, use_ai=False)

# 서버가 뜨자마자(또는 지난 세션이 만들어둔 sample_output/조립결과.xlsx가
# 디스크에 남아있는 채로) 2/3단계에 들어가면 실제로는 1단계를 실행한 적
# 없는데도 예시/이전 결과가 그대로 보이는 문제가 있었다(사용자 피드백) --
# "이번 서버 프로세스에서 /api/build를 한 번이라도 실제로 돌렸는지"를 별도로
# 추적해서, 프론트가 이 값으로 2/3단계 진입 자체를 막는다.
_build_completed = False


@app.get("/api/build-status")
def build_status():
    return {"built": _build_completed}


class DecisionRequest(BaseModel):
    status: str  # "approved" | "rejected"


@app.get("/api/items")
def list_items():
    return list(_ITEMS_STORE.values())


@app.post("/api/items/{item_id}/decision")
def decide(item_id: str, body: DecisionRequest):
    if item_id not in _ITEMS_STORE:
        raise HTTPException(404, "item not found")
    if body.status not in ("approved", "rejected"):
        raise HTTPException(400, "status must be approved or rejected")
    _ITEMS_STORE[item_id]["status"] = body.status
    return _ITEMS_STORE[item_id]


@app.get("/api/pbc-preview")
def pbc_preview(sheet: str):
    """수식이 참조하는 시트명으로 원본 PBC 데이터를 확정적으로(AI 검색 없이)
    찾아서 보여준다 -- 셀에 다른 시트를 가리키는 수식이 있을 때 씀."""
    path = pbc_library.find_by_sheet_name(SAMPLE_DATA_DIR, sheet)
    if path is None:
        raise HTTPException(404, "no matching PBC file for this sheet")
    preview = pbc_library.load_table_preview(path)
    return {"sheet": sheet, "sourcePath": path, "sourceName": Path(path).name, **preview}


_SYNTHETIC_PBC_DIR = (FIXTURES_DIR / "synthetic_pbc").resolve()


@app.get("/api/documents/{doc_name}")
def get_document(doc_name: str):
    """승인형 데모 항목(특수관계자거래/담보제공자산)의 후보 문서(PDF)를
    브라우저에서 바로 볼 수 있게 원본 파일 그대로 스트리밍한다 -- AI는
    찾기까지만 하고, 실제 내용 확인은 사람이 원본을 직접 보고 판단한다는
    원칙 그대로."""
    path = (_SYNTHETIC_PBC_DIR / doc_name).resolve()
    if _SYNTHETIC_PBC_DIR not in path.parents or not path.exists():
        raise HTTPException(404, "document not found")
    return FileResponse(str(path), media_type="application/pdf")


# ---- 왼쪽 패널: 1단계가 조립한 조서 엑셀(여러 시트) ------------------------
# build_workpaper()가 실제로 만드는 시트 전부를 sample_output/조립결과.xlsx에서
# 읽어 그대로 보여준다 -- 진짜 파이프라인 출력물을 검토하는 화면.
WORKPAPER_PATH = BUILD_OUTPUT_PATH
PBC_DIVIDER_SHEET = "PBC >>>>>"

WORKPAPER_SHEETS = [
    "기준가격대장(결산후)",
    "통합기준가격대장(결산후)",
    "C1_정산표",
    "C2_자산부채평가",
    "C3_수익",
    "평가액검증(주식)",
    "평가액검증(수익증권)",
    "평가액검증(선물)",
    "C4_수수료비용 등",
    "C5_F.N",
    "재무제표검증",
    PBC_DIVIDER_SHEET,
    "가중평균좌수",
    "국내유동명세",
    "국내주식명세",
    "펀드별명세",
    "수익증권명세",
    "선물명세",
    "채권명세",
    "보수분배내역",
    "일별자산내역",
    "설정해지내역",
    "KRS정보시스템_주식",
    "KRX정보시스템_ETF",
    "KRX정보시스템_코스피200",
]
SHEET_MAX_ROWS = 1000
SHEET_MAX_COLS = 60

_workpaper_wb = None  # data_only=True -- 계산된 값 (화면 표시용)
_workpaper_wb_formulas = None  # data_only=False -- 수식 원문 (셀 클릭 시 보여줄 용도)


def _get_workpaper_wb():
    global _workpaper_wb
    if _workpaper_wb is None:
        _workpaper_wb = openpyxl.load_workbook(str(WORKPAPER_PATH), data_only=True)
    return _workpaper_wb


def _get_workpaper_wb_formulas():
    global _workpaper_wb_formulas
    if _workpaper_wb_formulas is None:
        _workpaper_wb_formulas = openpyxl.load_workbook(str(WORKPAPER_PATH), data_only=False)
    return _workpaper_wb_formulas


def _find_caption_row(ws, caption_substring: str) -> int | None:
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row, 3).value
        if v and caption_substring in str(v):
            return row
    return None


# 특수관계자거래/담보제공자산 주석에 채울 "가공(데모) 데이터".
# 예전엔 빌드 시점에 시트에 바로 써넣었지만, 사용자 요청으로 "1단계 실행
# 직후엔 비어있고, 검토 화면에서 예(승인)를 눌러야 반영"되게 바꿨다. 그래서
# 빌드 때는 값을 시트에 쓰지 않고, 채웠을 때 보일 값만 메모리에 계산해둔다
# ({(엑셀행, 엑셀열): 표시문자열}). get_sheet가 해당 항목이 승인 상태일 때만
# 이 값을 C5_F.N 응답에 덮어씌운다("AI가 이 계약서를 찾아 이렇게 반영하면
# 될까요?" -> 사람이 예 -> 그제서야 시트에 반영, 이라는 흐름).
_TEUKSU_OVERLAY: dict[tuple[int, int], str] = {}
_DAMBO_OVERLAY: dict[tuple[int, int], str] = {}


def _round_k(v, div: int = 1) -> str:
    """원 단위 값을 천원 단위 정수 문자열로 (엑셀 ROUND(...,0)와 동일한 반올림).
    조서 전반의 '(단위: 천원)' 표기와 맞춤."""
    if v is None:
        return ""
    x = v / div / 1000
    n = int(x + 0.5) if x >= 0 else -int(-x + 0.5)
    return str(n)


def _compute_demo_overlays(path: Path) -> None:
    """빌드(+엑셀 재계산) 결과에서 특수관계자거래/담보제공자산 주석에 채울 값을
    미리 계산해 메모리에 담아둔다 -- 지어낸 숫자가 아니라 이미 조립된 워크북의
    실제 값에서 파생된다.

    특수관계자거래: 당기 보수는 C4_수수료비용 등의 PL보수금액, 전기 보수·당기/
    전기 미지급보수는 c5_bosu_note.py가 같은 4개 카테고리에 쓰는 것과 동일한
    원본('C1_정산표'!I/N/O, 행 110~113/56~59)에서 천원 환산. 판매회사가 2곳인데
    C4/C1엔 판매 카테고리가 1개뿐이라 반씩 나눈다(가상 데이터라 배분 기준은
    편의상 결정, 사용자 확인). 수탁회사(한아름은행)는 이 노트 구조에 행이
    없다 -- 수탁회사는 특수관계자가 아니라는 실제 파일의 판단.

    담보제공자산: 이 시트 '1. 지분증권' 노트가 이미 보여주는 보유종목 2개
    (종목명/수량/결산후 장부금액)를 그대로 담보 목록으로 쓴다."""
    _TEUKSU_OVERLAY.clear()
    _DAMBO_OVERLAY.clear()
    wb = openpyxl.load_workbook(str(path), data_only=True)
    if "C5_F.N" not in wb.sheetnames:
        return
    ws = wb["C5_F.N"]
    c1 = wb["C1_정산표"] if "C1_정산표" in wb.sheetnames else None
    c4 = wb["C4_수수료비용 등"] if "C4_수수료비용 등" in wb.sheetnames else None

    teuksu_row = _find_caption_row(ws, "특수관계자와의 거래")
    if teuksu_row is not None and c1 is not None and c4 is not None:
        fdr = teuksu_row + 5  # c5_teuksu_note.py의 first_data_row (header_row+1)
        n56, n57, n59 = c1["N56"].value, c1["N57"].value, c1["N59"].value
        o56, o57, o59 = c1["O56"].value, c1["O57"].value, c1["O59"].value
        i110, i111, i113 = c1["I110"].value, c1["I111"].value, c1["I113"].value
        d9, d10, d12 = c4["D9"].value, c4["D10"].value, c4["D12"].value
        # (offset, 당기=E열, 전기=F열) -- 기존 후처리 수식과 동일 소스.
        entries = [
            (0, _round_k(n56), _round_k(o56)),          # 자산운용 미지급
            (1, _round_k(d9), _round_k(i110)),          # 자산운용 보수
            (2, _round_k(n57, 2), _round_k(o57, 2)),    # 판매1 미지급
            (3, _round_k(d10, 2), _round_k(i111, 2)),   # 판매1 보수
            (4, _round_k(n57, 2), _round_k(o57, 2)),    # 판매2 미지급
            (5, _round_k(d10, 2), _round_k(i111, 2)),   # 판매2 보수
            (6, _round_k(n59), _round_k(o59)),          # 사무관리 미지급
            (7, _round_k(d12), _round_k(i113)),         # 사무관리 보수
        ]
        for off, e_val, f_val in entries:
            _TEUKSU_OVERLAY[(fdr + off, 5)] = e_val
            _TEUKSU_OVERLAY[(fdr + off, 6)] = f_val

    jibun_row = _find_caption_row(ws, "1. 지분증권")
    dambo_row = _find_caption_row(ws, "담보제공자산")
    if jibun_row is not None and dambo_row is not None:
        jfdr = jibun_row + 6   # c5_jibun_note.py first_data_row (header+2)
        dfdr = dambo_row + 6   # c5_dambo_note.py first_data_row (header+1)
        total_row = dambo_row + 9  # header + 1 + BUFFER_ROWS(3)
        total_qty = 0
        total_amt = 0
        for i in range(2):
            src = jfdr + i
            name = ws.cell(src, 3).value  # 종목명
            qty = ws.cell(src, 4).value   # 수량
            amt = ws.cell(src, 6).value   # 결산후 장부금액(천원)
            row = dfdr + i
            _DAMBO_OVERLAY[(row, 3)] = "지분증권"
            _DAMBO_OVERLAY[(row, 4)] = "" if name is None else str(name)
            _DAMBO_OVERLAY[(row, 5)] = "" if qty is None else str(int(qty))
            _DAMBO_OVERLAY[(row, 6)] = "" if amt is None else str(int(amt))
            total_qty += int(qty or 0)
            total_amt += int(amt or 0)
        _DAMBO_OVERLAY[(total_row, 5)] = str(total_qty)
        _DAMBO_OVERLAY[(total_row, 6)] = str(total_amt)


def _c5_overlay_for_status() -> dict[tuple[int, int], str]:
    """지금 승인된 데모 항목에 해당하는 오버레이만 모아서 돌려준다 -- C5_F.N을
    보여줄 때/다운로드할 때 이 값으로 빈 셀을 덮어씌운다."""
    overlay: dict[tuple[int, int], str] = {}
    teuksu = _ITEMS_STORE.get("demo::특수관계자거래")
    if teuksu and teuksu.get("status") == "approved":
        overlay.update(_TEUKSU_OVERLAY)
    dambo = _ITEMS_STORE.get("demo::담보제공자산")
    if dambo and dambo.get("status") == "approved":
        overlay.update(_DAMBO_OVERLAY)
    return overlay


def _add_pbc_divider_sheet(path: Path) -> None:
    """"재무제표검증"(마지막 조립 시트)과 "가중평균좌수"(첫 원본 PBC 시트)
    사이에 제목만 있는 빈 시트를 끼워 넣어, 탭 목록에서 "여기부터는 조립된
    조서가 아니라 원본 PBC 자료다"를 한눈에 구분할 수 있게 한다 -- 순수
    화면 표시용 구분선이라 build_workpaper() 자체(core/assemble.py)를 건드릴
    필요 없이 여기(api.py)서 후처리로 끼워 넣는다."""
    wb = openpyxl.load_workbook(str(path))
    if PBC_DIVIDER_SHEET in wb.sheetnames:
        return
    anchor = "재무제표검증"
    index = wb.sheetnames.index(anchor) + 1 if anchor in wb.sheetnames else len(wb.sheetnames)
    wb.create_sheet(PBC_DIVIDER_SHEET, index)
    wb.save(str(path))


def _strip_xlsx_metadata(path: Path) -> None:
    """엑셀 문서 속성의 작성자/최종수정자/회사명을 비운다 -- Excel로 재계산해
    저장하면 그 환경 사용자 이름이 lastModifiedBy에 박히므로, 배포용 산출물에
    개인 이름이 남지 않도록 후처리한다. openpyxl로 다시 저장하면 캐시된 수식
    값이 날아가므로, zip 안의 docProps/*.xml만 직접 손봐서 시트 데이터는 그대로
    둔다."""
    import re
    import zipfile

    tmp = path.with_suffix(path.suffix + ".tmp")
    try:
        with zipfile.ZipFile(str(path), "r") as zin, zipfile.ZipFile(str(tmp), "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "docProps/core.xml":
                    t = data.decode("utf-8")
                    t = re.sub(r"<dc:creator>.*?</dc:creator>", "<dc:creator></dc:creator>", t, flags=re.S)
                    t = re.sub(
                        r"<cp:lastModifiedBy>.*?</cp:lastModifiedBy>", "<cp:lastModifiedBy></cp:lastModifiedBy>", t, flags=re.S
                    )
                    data = t.encode("utf-8")
                elif item.filename == "docProps/app.xml":
                    t = data.decode("utf-8")
                    t = re.sub(r"<Company>.*?</Company>", "<Company></Company>", t, flags=re.S)
                    t = re.sub(r"<Manager>.*?</Manager>", "<Manager></Manager>", t, flags=re.S)
                    data = t.encode("utf-8")
                zout.writestr(item, data)
        os.replace(str(tmp), str(path))
    except Exception:  # noqa: BLE001
        if tmp.exists():
            tmp.unlink()


def _excel_com_available() -> bool:
    """이 환경에서 Excel COM 재계산(_recalculate_with_excel)이 가능한지.
    Windows + pywin32 + 로컬 Excel이 있어야 한다. 리눅스 배포(Render 등)에서는
    False -- 이때는 커밋된 sample_output/조립결과.xlsx(값까지 캐싱된 완성본)를
    다시 만들지 않고 그대로 서빙한다. DISABLE_EXCEL_COM=1 로 명시적으로 끌 수도
    있다(리눅스 배포 흐름을 윈도우에서 테스트하거나, 로컬에서도 pre-built
    데모로 강제하고 싶을 때)."""
    if os.environ.get("DISABLE_EXCEL_COM") == "1":
        return False
    try:
        import win32com.client  # noqa: F401
    except ImportError:
        return False
    return True


_PREBUILT_REPORT_PATH = PROJECT_ROOT / "sample_output" / "build_report.json"


def _load_prebuilt_report() -> dict:
    """no-COM(배포) 경로에서 build_workpaper()를 돌리는 대신 읽는, 표본 조립의
    사전 계산된 report(신규계정/드리프트 목록). 로컬에서 실제 표본을 조립해 뽑아
    커밋해둔 파일이라 값이 실제 산출물과 일치한다. 없으면 빈 report로 폴백."""
    import json

    try:
        return json.loads(_PREBUILT_REPORT_PATH.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return {"missing_account_codes": [], "drifted_prior_year_rows": []}


def _recalculate_with_excel(path: Path) -> None:
    """build_workpaper()는 수식만 써넣고 계산된 값은 저장하지 않는다 -- "엑셀이
    열 때 알아서 재계산한다"는 전제로 만들어진 설계라, 실제 엑셀에서 열면
    문제 없지만 openpyxl로 캐시값만 읽는 이 앱은 모든 수식 셀이 빈 값으로
    보인다 (라벨/캡션은 리터럴이라 괜찮은데 숫자 컬럼은 전부 이 문제).
    build_workpaper() 쪽을 고칠 문제가 아니라(그쪽은 원래 계획대로 잘 동작),
    "캐시값만 읽는" 이 앱의 표시 요구사항이라 여기서 후처리한다: 실제 엑셀을
    백그라운드로 띄워 한 번 열었다 재계산 후 저장한다. pywin32 필요, Windows +
    로컬 엑셀 설치 전제 -- 실패해도(엑셀 없는 환경 등) 빌드 자체는 이미
    끝났으니 조용히 넘어간다(수식 텍스트 기반 근거 연결 기능은 이 값 없이도
    동작한다)."""
    try:
        import pythoncom
        import win32com.client
    except ImportError:
        return
    try:
        # FastAPI가 이 함수를 스레드풀 워커 스레드에서 돌리는데, COM은 스레드마다
        # CoInitialize를 따로 호출해줘야 한다 -- 안 하면 DispatchEx가
        # "CoInitialize가 호출되지 않았습니다" 에러를 던진다 (메인 스레드에서
        # 직접 실행할 땐 암묵적으로 되어 있어서 이 문제가 안 보였다).
        pythoncom.CoInitialize()
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            try:
                wb = excel.Workbooks.Open(str(path))
                excel.CalculateFullRebuild()
                wb.Save()
                wb.Close(SaveChanges=False)
            finally:
                excel.Quit()
        finally:
            pythoncom.CoUninitialize()
    except Exception:
        # 엑셀 미설치 등 이 앱 표시 요구사항과 무관한 환경에서는 조용히
        # 넘어가되(빌드 자체는 이미 끝났으므로), 무슨 예외였는지는 남겨서
        # 다음에 같은 문제(예: 위 CoInitialize 문제)가 나면 로그로 보이게 한다.
        import traceback

        traceback.print_exc()


def _reset_workpaper_cache() -> None:
    """/api/build로 새로 조립한 뒤, 다음 조회부터 새 파일을 다시 읽도록
    캐시를 비운다."""
    global _workpaper_wb, _workpaper_wb_formulas
    _workpaper_wb = None
    _workpaper_wb_formulas = None


@app.post("/api/build")
def run_build(
    gijunkagyeok: UploadFile | None = File(None),
    gungnaeyudong: UploadFile | None = File(None),
    gungnaejusik: UploadFile | None = File(None),
    pundbyeolmyeongse: UploadFile | None = File(None),
    sooikjeunggwon: UploadFile | None = File(None),
    seonmul: UploadFile | None = File(None),
    chaegwon: UploadFile | None = File(None),
    gajungpyeonggyunjwasu: UploadFile | None = File(None),
    bosubunbae: UploadFile | None = File(None),
    ilbyeoljasan: UploadFile | None = File(None),
    seoljeongheji: UploadFile | None = File(None),
    reference_workpaper: UploadFile | None = File(None),
):
    """1단계: PBC 원본자료를 실제로 받아서 build_workpaper()를 돌리고, 그
    결과를 2단계가 바로 볼 수 있게 sample_output/조립결과.xlsx에 저장한다.
    필드를 비우면 sample_data/의 표본으로 대신 채운다 -- 전체 흐름을 빠르게
    확인하고 싶을 때와 실제 파일을 올릴 때를 같은 엔드포인트로 처리.

    def(async 아님)로 선언해 FastAPI가 스레드풀에서 돌리게 한다 --
    build_workpaper()가 몇 초 걸릴 수 있는 동기 작업이라, 이벤트 루프를
    막지 않기 위함."""
    uploads = {
        "gijunkagyeok": gijunkagyeok,
        "gungnaeyudong": gungnaeyudong,
        "gungnaejusik": gungnaejusik,
        "pundbyeolmyeongse": pundbyeolmyeongse,
        "sooikjeunggwon": sooikjeunggwon,
        "seonmul": seonmul,
        "chaegwon": chaegwon,
        "gajungpyeonggyunjwasu": gajungpyeonggyunjwasu,
        "bosubunbae": bosubunbae,
        "ilbyeoljasan": ilbyeoljasan,
        "seoljeongheji": seoljeongheji,
        "reference_workpaper": reference_workpaper,
    }

    BUILD_OUTPUT_PATH.parent.mkdir(exist_ok=True)
    # Excel COM 재계산이 가능한 환경(Windows+Excel)에서만 실제로 조립본을 새로
    # 써서 값을 채운다. 리눅스 배포처럼 COM이 없는 곳에서는 (1) build_workpaper가
    # 수식만 넣고 값은 못 채우고, (2) 12개 CSV를 pandas로 읽어 24시트 워크북을
    # 메모리에 만드는 조립 자체가 무거워 무료 인스턴스(512MB)에서 OOM으로 죽는다.
    # 그래서 no-COM 경로에서는 조립을 아예 돌리지 않고, 커밋된 pre-built 완성본
    # (값까지 캐싱됨)을 그대로 서빙하고 report는 사전 계산된 JSON을 읽는다 --
    # 화면에 보이는 데이터는 어차피 이 완성본이고, 검토 항목(특수관계자거래/
    # 담보제공자산)은 report가 아니라 데모 매니페스트에서 나온다.
    com_available = _excel_com_available()
    if com_available:
        with tempfile.TemporaryDirectory() as tmp_str:
            tmp = Path(tmp_str)
            paths: dict[str, Path] = {}
            for field, sample_filename in SAMPLE_BUILD_INPUTS.items():
                upload = uploads[field]
                if upload is not None:
                    dest = tmp / upload.filename
                    with open(dest, "wb") as out:
                        shutil.copyfileobj(upload.file, out)
                    paths[field] = dest
                else:
                    paths[field] = SAMPLE_DATA_DIR / sample_filename
            try:
                report = build_workpaper(
                    paths["gijunkagyeok"],
                    paths["gungnaeyudong"],
                    paths["gungnaejusik"],
                    paths["pundbyeolmyeongse"],
                    paths["sooikjeunggwon"],
                    paths["seonmul"],
                    paths["chaegwon"],
                    paths["gajungpyeonggyunjwasu"],
                    paths["bosubunbae"],
                    paths["ilbyeoljasan"],
                    paths["seoljeongheji"],
                    paths["reference_workpaper"],
                    BUILD_OUTPUT_PATH,
                )
            except Exception as e:  # noqa: BLE001
                raise HTTPException(500, f"조립 실패: {e}") from e
        # 방금 새로 조립한 파일을 후처리(구분시트 추가 → 값 재계산 → 작성자명 제거).
        _add_pbc_divider_sheet(BUILD_OUTPUT_PATH)
        _recalculate_with_excel(BUILD_OUTPUT_PATH)
        _strip_xlsx_metadata(BUILD_OUTPUT_PATH)  # Excel 재계산이 박은 작성자명 제거
    else:
        # 커밋된 완성본은 이미 구분시트+값+메타정리가 끝난 상태라 손대지 않고,
        # report는 같은 표본을 로컬에서 조립해 미리 뽑아둔 JSON을 읽는다.
        report = _load_prebuilt_report()
    _reset_workpaper_cache()
    _init_store(report, use_ai=USE_LIVE_AI)
    # 특수관계자거래/담보제공자산 값은 시트에 쓰지 않고 오버레이로만 계산해둔다
    # -- 검토 화면에서 예(승인)를 눌러야 get_sheet가 이 값을 덮어씌운다.
    _compute_demo_overlays(BUILD_OUTPUT_PATH)
    global _build_completed
    _build_completed = True
    return {
        "ok": True,
        "missingAccountCount": len(report.get("missing_account_codes", [])),
        "usedSampleFallback": [f for f, u in uploads.items() if u is None],
    }


@app.get("/api/workpaper/sheets")
def list_sheets():
    wb = _get_workpaper_wb()
    return [name for name in WORKPAPER_SHEETS if name in wb.sheetnames]


@app.get("/api/workpaper/sheets/{name}")
def get_sheet(name: str):
    wb_v = _get_workpaper_wb()
    wb_f = _get_workpaper_wb_formulas()
    if name not in WORKPAPER_SHEETS or name not in wb_v.sheetnames:
        raise HTTPException(404, "sheet not found")
    ws_v = wb_v[name]
    ws_f = wb_f[name]
    max_row = min(ws_v.max_row, SHEET_MAX_ROWS)
    max_col = min(ws_v.max_column, SHEET_MAX_COLS)
    rows = []
    for r in range(1, max_row + 1):
        row = []
        for c in range(1, max_col + 1):
            v = ws_v.cell(r, c).value
            fv = ws_f.cell(r, c).value
            formula = fv if isinstance(fv, str) and fv.startswith("=") else None
            row.append({"v": "" if v is None else str(v), "f": formula})
        rows.append(row)
    # C5_F.N: 특수관계자거래/담보제공자산이 승인 상태면, 빈 채로 조립된 셀에
    # 미리 계산해둔 가공 데이터를 덮어씌운다(검토 화면에서 예를 눌러 반영한 것).
    if name == "C5_F.N":
        for (rr, cc), val in _c5_overlay_for_status().items():
            if 1 <= rr <= len(rows) and 1 <= cc <= len(rows[rr - 1]):
                rows[rr - 1][cc - 1]["v"] = val
    return {
        "name": name,
        "rows": rows,
        "truncated": ws_v.max_row > SHEET_MAX_ROWS or ws_v.max_column > SHEET_MAX_COLS,
    }


@app.get("/api/prior-year-financials")
def prior_year_financials(section: str):
    """C1_정산표의 재무상태표/손익계산서 구간을 클릭했을 때 보여줄 근거.
    이 시트에는 원본 PBC 파일이 따로 없다 -- 여기서 하는 작업 자체가 전기
    재무제표 데이터를 손으로 입력하는 것이라, "근거"는 core.prior_year_source가
    같은 참조 워크북의 전기(I/J) 컬럼에서 재구성한 값 그 자체다."""
    if section not in ("재무상태표", "손익계산서"):
        raise HTTPException(400, "section must be 재무상태표 or 손익계산서")
    wb = _get_workpaper_wb()
    items = [i for i in prior_year_source.extract_prior_year_financials(wb) if i.section == section]
    return {
        "section": section,
        "items": [{"label": i.label, "amount": i.amount, "row": i.row} for i in items],
    }


@app.get("/api/prior-year-capital-statement")
def prior_year_capital_statement():
    """C1_정산표의 <C.E>(자본변동표) 구간용. 원리는 prior_year_financials와
    같다 -- 같은 참조 워크북의 전기 자본변동 내역(설정/해지/손익/조정)을
    core.prior_year_source가 재구성한 값 그대로 보여준다."""
    wb = _get_workpaper_wb()
    rows = prior_year_source.extract_prior_year_capital_statement(wb)
    return {
        "items": [
            {"label": r.label, "principal": r.principal, "retainedEarnings": r.retained_earnings, "total": r.total}
            for r in rows
        ]
    }


@app.get("/api/workpaper/download")
def download_workpaper():
    """3단계(완료) 다운로드 버튼용. build_workpaper()를 실제로 돌리지 않으므로
    (core/assemble.py는 다른 세션에서 편집 중), 왼쪽 패널에서 이미 보여주고
    있는 WORKPAPER_SHEETS 15개를 참조 워크북에서 그대로 복사해 새 파일로
    묶어 낸다 -- "이 도구가 검토 대상으로 다루는 시트"만 담은 데모용 산출물.
    수식은 그대로 유지되고(copy_sheet), 이 15개 시트끼리만 서로 참조하므로
    엑셀에서 열어도 깨지지 않는다."""
    src_wb = _get_workpaper_wb_formulas()
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)
    for name in WORKPAPER_SHEETS:
        if name in src_wb.sheetnames:
            copy_sheet(src_wb, name, out_wb, unhide=True)

    # 화면에서 승인(예)한 특수관계자거래/담보제공자산 값은 다운로드본에도
    # 담는다 -- 리터럴로 써넣으므로(엑셀 재계산 불필요) 그대로 보인다.
    overlay = _c5_overlay_for_status()
    if overlay and "C5_F.N" in out_wb.sheetnames:
        out_c5 = out_wb["C5_F.N"]
        for (rr, cc), val in overlay.items():
            out_c5.cell(rr, cc).value = int(val) if val.isdigit() else val

    buf = io.BytesIO()
    out_wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        # Content-Disposition 헤더는 latin-1만 허용돼서 한글 파일명은 못 씀 --
        # RFC 5987 filename*로 UTF-8 원본 파일명을 같이 실어 보낸다 (지원
        # 브라우저는 이걸 쓰고, 아니면 ascii filename으로 폴백).
        headers={
            "Content-Disposition": (
                "attachment; filename=\"workpaper_demo.xlsx\"; "
                "filename*=UTF-8''%EA%B2%80%ED%86%A0%EC%A1%B0%EC%84%9C_%EB%8D%B0%EB%AA%A8.xlsx"
            )
        },
    )


# ---- 프론트엔드(React) 정적 서빙 -----------------------------------------
# 배포 시 프론트와 백엔드를 한 서비스로 합친다: Vite로 빌드한 frontend/dist를
# 이 FastAPI 앱이 직접 서빙하면 URL 하나·CORS 불필요·관리 지점 하나로 끝난다.
# 로컬 개발(vite dev로 5173에서 따로 띄우는 경우)에는 dist가 없을 수 있으니,
# 존재할 때만 마운트한다. 이 블록은 모든 /api 라우트가 등록된 뒤(파일 맨 끝)
# 실행되어야 catch-all이 API를 가리지 않는다.
_FRONTEND_DIST = PROJECT_ROOT / "frontend" / "dist"
if _FRONTEND_DIST.is_dir():
    _INDEX_HTML = _FRONTEND_DIST / "index.html"

    # 해시된 정적 자산(js/css/이미지)은 /assets 아래에 있다.
    app.mount(
        "/assets",
        StaticFiles(directory=str(_FRONTEND_DIST / "assets")),
        name="assets",
    )

    @app.get("/{full_path:path}")
    def _spa_fallback(full_path: str):
        """SPA 라우팅용 catch-all: /api로 시작하지 않는 모든 경로는 index.html을
        돌려줘 클라이언트 라우터(/review, /complete 등)가 처리하게 한다. 실제
        파일(favicon 등)이 dist에 있으면 그 파일을 우선 준다."""
        if full_path.startswith("api/"):
            raise HTTPException(404, "not found")
        candidate = (_FRONTEND_DIST / full_path).resolve()
        if _FRONTEND_DIST in candidate.parents and candidate.is_file():
            return FileResponse(str(candidate))
        return FileResponse(str(_INDEX_HTML))
